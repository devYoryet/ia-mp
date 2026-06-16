"""Panel del Clasificador IA — FastAPI.

Vistas:
- /            Resumen
- /comparacion Backtest (IA vs personas) — gasto del backtest Y acumulado
- /revision    Cola de revisión paginada y filtrable. Cada fila marcada como
               INTERÉS / DESCARTE / PACTIVO NUEVO. El pactivo se elige del
               catálogo y la composición/presentación se filtran por pactivo
               (igual que el legacy de gestor_licitaciones). Aprobar la hoja,
               corregir o descartar una línea (con motivo obligatorio).
- /reglas      Reglas de negocio y correcciones (feedback al prompt)
- /legacy      Módulos portados desde gestor_oc Laravel (Subida TD,
               Importaciones, Adjudicaciones, Item Detalle). Cada uno recibe
               un archivo por chunks y lanza un script de bin/ en background.
- /api/catalogo  comp/pres válidas de un pactivo (para los selects dependientes)

Servir detrás de nginx en https://iabot.pharmatender.cl
    uvicorn api.main:app --host 0.0.0.0 --port 8800
"""

from __future__ import annotations

import html
import pathlib
import sys
import time
from datetime import datetime

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from starlette.middleware.sessions import SessionMiddleware

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))
from config import config  # noqa: E402
from db import conectar  # noqa: E402
from descarte_modelo import cargar_modelo_descarte, prob_descarte  # noqa: E402
from reglas import normalizar  # noqa: E402

from api.auth import router as auth_router, usuario_actual  # noqa: E402
from api.ui import EMAILS_PACTIVOS_EXTRA  # noqa: E402
from api.legacy import router as legacy_router  # noqa: E402
from api.ui import layout as _layout  # noqa: E402

app = FastAPI(title="Clasificador IA — Pharmatender")
app.include_router(auth_router)
app.include_router(legacy_router)


# Rutas que NO requieren login. /salud queda público para el healthcheck del
# container; /api/catalogo se usa por el front del propio panel (ya logueado).
_RUTAS_PUBLICAS = {"/login", "/logout", "/salud"}

# Vista TÉCNICA (badges de confianza/vía/entrenamiento, de dónde viene el error):
# solo para quien afina el sistema. El resto ve la vista LIMPIA de aprobación.
_EMAIL_ADMIN = "y.danoun@pharmatender.cl"
# Quiénes pueden ver el campo de GASTOS (costo acumulado) en /resumen.
_EMAILS_GASTOS = {
    "y.danoun@pharmatender.cl",
    "m.moraga@pharmatender.cl",
    "m.saavedra@pharmatender.cl",
}


def _norm_fecha(v: str, fin: bool = False) -> "str | None":
    """Normaliza un filtro de fecha del panel. Acepta 'YYYY-MM-DD' (día completo)
    o 'YYYY-MM-DDTHH:MM' / 'YYYY-MM-DD HH:MM' (datetime-local, corte por hora).
    Devuelve 'YYYY-MM-DD HH:MM:SS' o None. `fin=True` extiende al final del rango."""
    v = (v or "").strip().replace("T", " ")
    if not v:
        return None
    if len(v) <= 10:  # solo fecha
        return v + (" 23:59:59" if fin else " 00:00:00")
    if len(v) == 16:  # fecha + HH:MM
        return v + (":59" if fin else ":00")
    return v


def _es_admin(usuario: dict | None) -> bool:
    return bool(usuario) and (usuario.get("email") or "").strip().lower() == _EMAIL_ADMIN


def _ve_gastos(usuario: dict | None) -> bool:
    return bool(usuario) and (usuario.get("email") or "").strip().lower() in _EMAILS_GASTOS


@app.middleware("http")
async def proteger(request: Request, call_next):
    """Si no hay sesión, redirige a /login conservando la URL pretendida."""
    ruta = request.url.path
    if ruta in _RUTAS_PUBLICAS:
        return await call_next(request)
    if usuario_actual(request) is None:
        # Conservar el destino solo para GETs (un POST sin sesión queda perdido
        # de todos modos y reintentarlo con next= sería confuso).
        if request.method == "GET":
            destino = ruta + (f"?{request.url.query}" if request.url.query else "")
            return RedirectResponse(f"/login?next={destino}", status_code=303)
        return RedirectResponse("/login", status_code=303)
    return await call_next(request)


# SessionMiddleware se agrega ÚLTIMO: en Starlette, el último middleware
# agregado es el más externo. Necesitamos que envuelva a `proteger` para que
# cuando `proteger` lea `request.session`, la cookie ya esté decodificada.
# Cookie firmada (HMAC + itsdangerous), 8h de vida — alcanza para una jornada.
app.add_middleware(
    SessionMiddleware,
    secret_key=config.session_secret,
    session_cookie="iabot_sess",
    max_age=8 * 3600,
    same_site="lax",
    https_only=False,  # nginx termina TLS; la cookie viaja en HTTP interno.
)

TABLAS_VALIDAS = ("compra_agil", "Licitaciones_diarias")
POR_HOJA_DEFAULT = 50  # filas por hoja en la cola de revisión (configurable)
POR_HOJA_OPCIONES = (25, 50, 100, 200)

def _query(sql: str, args=()) -> list:
    conn = conectar()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, args)
            return list(cur.fetchall())
    finally:
        conn.close()


def _e(v) -> str:
    return html.escape("" if v is None else str(v))


# Modelo de descarte entrenado — se carga una vez para mostrar, por fila, la
# etiqueta de "entrenamiento": qué dice el clasificador entrenado (interés o
# descarte), independiente de qué etapa de la cascada resolvió la fila.
_MODELO_DESCARTE = None
_MODELO_CARGADO = False


def _modelo_descarte():
    global _MODELO_DESCARTE, _MODELO_CARGADO
    if not _MODELO_CARGADO:
        try:
            _MODELO_DESCARTE = cargar_modelo_descarte()
        except Exception:  # noqa: BLE001
            _MODELO_DESCARTE = None
        _MODELO_CARGADO = True
    return _MODELO_DESCARTE


# ----------------------------------------------------------------- Catálogo ---
# Catálogo de clasificación EN MEMORIA: {pactivo_normalizado: {nombre, comp, pres}}.
# Mismo origen que el legacy de gestor_licitaciones: 0001_td_oc.Base ∪ la tabla
# `diccionario`. Los selects dependientes pactivo→comp→pres salen de acá sin
# tocar la BD en cada interacción.
#
# El catálogo NO es fijo: CRECE a medida que se ingresan pactivos/composiciones/
# presentaciones nuevos. Por eso se refresca solo cada `_CATALOGO_TTL` — un panel
# de larga vida toma las altas nuevas sin reiniciar.
_CATALOGO_TTL = 8 * 3600  # segundos (8 horas)
_CATALOGO_CACHE = None
_CATALOGO_TS = 0.0


def _catalogo(forzar: bool = False) -> dict:
    global _CATALOGO_CACHE, _CATALOGO_TS
    if (not forzar and _CATALOGO_CACHE is not None
            and time.time() - _CATALOGO_TS < _CATALOGO_TTL):
        return _CATALOGO_CACHE
    nombre, comp, pres = {}, {}, {}
    conn = conectar()
    try:
        with conn.cursor() as cur:
            cur.execute(
                f"SELECT DISTINCT Pactivo p, Comp c, MedidaPHT m "
                f"FROM `{config.db_catalogo}`.Base "
                f"WHERE Pactivo IS NOT NULL AND Pactivo<>''"
            )
            filas = list(cur.fetchall())
            cur.execute(
                f"SELECT DISTINCT pactivo p, comp c, presentacion m "
                f"FROM `{config.db_diccionario}`.diccionario "
                f"WHERE pactivo IS NOT NULL AND pactivo<>''"
            )
            filas += list(cur.fetchall())
            # pactivos_extra: catálogo manual (Carolina/admin). No traen comp/pres
            # propias —se completan al aprobar la 1ª fila—, pero DEBEN aparecer en
            # el autocompletado y dar encontrado:True para poder usarse en una fila.
            # Sin esto, un pactivo recién agregado al CRUD no se puede asignar.
            try:
                cur.execute(
                    "SELECT pactivo p, '' c, '' m FROM pactivos_extra WHERE activo=1"
                )
                filas += list(cur.fetchall())
            except Exception:  # noqa: BLE001
                pass  # entornos sin la tabla todavía
    finally:
        conn.close()
    for r in filas:
        p = (r["p"] or "").strip()
        if not p:
            continue
        k = normalizar(p)
        nombre.setdefault(k, p)
        if r["c"] and r["c"].strip():
            comp.setdefault(k, set()).add(r["c"].strip())
        if r["m"] and r["m"].strip():
            pres.setdefault(k, set()).add(r["m"].strip())
    _CATALOGO_CACHE = {
        k: {
            "nombre": nombre[k],
            "comp": sorted(comp.get(k, set())),
            "pres": sorted(pres.get(k, set())),
        }
        for k in nombre
    }
    _CATALOGO_TS = time.time()
    return _CATALOGO_CACHE


@app.get("/api/catalogo")
def api_catalogo(pactivo: str = ""):
    """Composiciones y presentaciones válidas de un pactivo (selects dependientes)."""
    info = _catalogo().get(normalizar(pactivo))
    if not info:
        return {"comp": [], "pres": [], "encontrado": False}
    return {"comp": info["comp"], "pres": info["pres"], "encontrado": True}


def _select(nombre: str, fila_n: int, clase: str, valor: str, opciones: list) -> str:
    """<select> con un valor actual + 'Sin Cla' + las opciones del catálogo."""
    valor = (valor or "").strip()
    vals: list[str] = []
    for v in [valor, "Sin Cla", *opciones]:
        v = (v or "").strip()
        if v and v not in vals:
            vals.append(v)
    ops = "".join(
        f"<option value=\"{_e(v)}\"{' selected' if v == valor else ''}>{_e(v)}</option>"
        for v in vals
    )
    return (
        f"<select name={nombre} class='{clase}' data-row='{fila_n}' "
        f"data-sug=\"{_e(valor)}\">{ops}</select>"
    )


# ---------------------------------------------------------------- Resumen ---
@app.get("/", response_class=HTMLResponse)
def resumen(request: Request):
    usuario = usuario_actual(request)
    # Los aprobadores no ven el resumen (gastos/métricas): van directo a la cola.
    if not _ve_gastos(usuario):
        return RedirectResponse("/revision", status_code=303)
    try:
        log = _query(
            "SELECT COUNT(*) n, SUM(revisado=0) pend, SUM(revisado=1) rev, "
            "SUM(revisado=1 AND feedback_correcto=1) ok FROM clasificador_ia_log"
        )[0]
        # El gasto real (todo lo cobrado por la API) sale del libro de costos,
        # que no se trunca entre backtests — a diferencia del log de clasificación.
        costo = _query(
            "SELECT IFNULL(SUM(costo_usd),0) c FROM clasificador_ia_costos"
        )[0]["c"]
        # Pendientes DE PROCESAR: filas nuevas sin clasificar (estado_gestor NULL
        # y sin clasificador) — lo que el worker tiene por delante.
        pend_proc = {}
        for _t in TABLAS_VALIDAS:
            pend_proc[_t] = _query(
                f"SELECT COUNT(*) n FROM `{_t}` WHERE estado_gestor IS NULL "
                f"AND (nombre_clasificador IS NULL OR nombre_clasificador='')"
            )[0]["n"]
    except Exception as exc:  # noqa: BLE001
        return _layout(
            "Error",
            f"<div class=vacio>No se pudo leer la base.<br><small>{_e(exc)}</small><br><br>"
            "¿Creaste las tablas con <code>schema/auditoria.sql</code>?</div>",
            usuario=usuario,
        )
    rev = log["rev"] or 0
    prec = f"{(log['ok'] or 0) / rev * 100:.1f}%" if rev else "—"
    # El card de GASTOS solo lo ven los autorizados (gerencia + quien afina).
    card_gastos = (
        f"<div class=card><div class=n>${float(costo):.2f}</div>"
        f"<div class=l>Gastado de ${config.budget_usd:.0f} (acumulado)</div></div>"
        if _ve_gastos(usuario) else ""
    )
    cuerpo = (
        "<h1>Resumen</h1>"
        "<h2>Pendientes de procesar (filas nuevas sin clasificar)</h2><div class=cards>"
        f"<div class=card><div class=n>{pend_proc['compra_agil']:,}</div>"
        f"<div class=l>Pendientes compra ágil</div></div>"
        f"<div class=card><div class=n>{pend_proc['Licitaciones_diarias']:,}</div>"
        f"<div class=l>Pendientes licitaciones</div></div>"
        "</div>"
        "<h2>Clasificación IA</h2><div class=cards>"
        f"<div class=card><div class=n>{log['n']}</div><div class=l>Clasificadas por IA</div></div>"
        f"<div class=card><div class=n>{log['pend'] or 0}</div><div class=l>Pendientes de revisión</div></div>"
        f"<div class=card><div class=n>{rev}</div><div class=l>Revisadas</div></div>"
        f"<div class=card><div class=n>{prec}</div><div class=l>Precisión (revisadas)</div></div>"
        f"{card_gastos}"
        "</div>"
        "<div class=aviso>Modo actual del worker: <b>" + _e(config.modo) + "</b>. "
        "La cola de revisión se llena cuando el worker corre en modo producción.</div>"
    )
    return _layout("Resumen", cuerpo, usuario=usuario)


# ------------------------------------------------------------ Comparación ---
# Cross matrix (humano × IA). Las 6 celdas + sus condiciones SQL. Click en cualquiera
# de las cards lleva a la lista de filas que cayeron ahí — para auditar puntualmente.
# Notación: hX_iY = humano=X, ia=Y. _NUEVO es ia.pactivo_nuevo IS NOT NULL.
_CELDAS = {
    # name : (etiqueta, color, SQL condition)
    "h1_i1":   ("✓ Acuerdo INTERÉS",    "ok",   "humano_estado_gestor=1 AND ia_interes=1 AND (ia_pactivo_nuevo IS NULL OR ia_pactivo_nuevo='')"),
    "h1_i0":   ("✗ FALSO NEGATIVO",      "fn",   "humano_estado_gestor=1 AND ia_interes=0"),
    "h1_nuevo":("⚠ Humano interés, IA nuevo", "warn", "humano_estado_gestor=1 AND ia_pactivo_nuevo IS NOT NULL AND ia_pactivo_nuevo<>''"),
    "h0_i1":   ("✗ FALSO POSITIVO",      "fp",   "humano_estado_gestor=0 AND ia_interes=1 AND (ia_pactivo_nuevo IS NULL OR ia_pactivo_nuevo='')"),
    "h0_i0":   ("✓ Acuerdo DESCARTE",    "ok",   "humano_estado_gestor=0 AND ia_interes=0"),
    "h0_nuevo":("⚠ Humano descarte, IA nuevo", "warn", "humano_estado_gestor=0 AND ia_pactivo_nuevo IS NOT NULL AND ia_pactivo_nuevo<>''"),
}


@app.get("/comparacion", response_class=HTMLResponse)
def comparacion(request: Request, cell: str = "", metodo: str = "",
                pactivo_sug: str = "", limit: int = 100) -> str:
    usuario = usuario_actual(request)
    # Si vienen filtros de drill-down, mostrar listado en vez de dashboard.
    if cell or metodo or pactivo_sug:
        return _comparacion_listado(cell, metodo, pactivo_sug, limit, usuario)
    return _comparacion_dashboard(usuario)


def _comparacion_dashboard(usuario: dict | None = None) -> str:
    try:
        tot = _query("SELECT COUNT(*) n FROM clasificador_ia_backtest")[0]["n"]
        if not tot:
            return _layout(
                "Backtest",
                "<h1>Backtest · IA vs personas</h1><div class=vacio>Aún no hay filas "
                "comparadas. El container <code>backtest</code> está procesando 200 "
                "filas cada 15 min — volvé en un rato.</div>",
                usuario=usuario,
            )
        r = _query(
            "SELECT COUNT(*) n, SUM(coincide_interes) ci, "
            "SUM(coincide_pactivo) cp, COUNT(coincide_pactivo) ncp, "
            "SUM(coincide_composicion) cc, SUM(coincide_presentacion) cpr, "
            "IFNULL(SUM(costo_usd),0) costo, IFNULL(AVG(costo_usd),0) prom "
            "FROM clasificador_ia_backtest"
        )[0]
        # Cross matrix: conteos de cada celda en una sola query (CASE WHEN ... THEN 1)
        case_parts = ", ".join(
            f"SUM(CASE WHEN {sql} THEN 1 ELSE 0 END) AS {name}"
            for name, (_, _, sql) in _CELDAS.items()
        )
        matriz = _query(f"SELECT {case_parts} FROM clasificador_ia_backtest")[0]
        # Por método — cuánto resuelve cada etapa, accuracy de pact, costo
        por_metodo = _query(
            "SELECT ia_metodo, COUNT(*) n, "
            "SUM(coincide_interes) ci, "
            "SUM(coincide_pactivo) cp, COUNT(coincide_pactivo) ncp, "
            "IFNULL(SUM(costo_usd),0) costo "
            "FROM clasificador_ia_backtest "
            "GROUP BY ia_metodo ORDER BY n DESC"
        )
        # Top pactivos sospechosos: IA sugirió X, humano descartó. Min N=5 para no
        # pescar ruido de clases con 1-2 filas.
        sospechosos = _query(
            "SELECT ia_pactivo p, "
            "COUNT(*) n, "
            "SUM(humano_estado_gestor=0) descartes, "
            "ROUND(SUM(humano_estado_gestor=0)/COUNT(*)*100,1) pct_desc "
            "FROM clasificador_ia_backtest "
            "WHERE ia_interes=1 AND ia_pactivo IS NOT NULL "
            "GROUP BY ia_pactivo HAVING n >= 5 AND pct_desc >= 50 "
            "ORDER BY pct_desc DESC, n DESC LIMIT 25"
        )
        acum = _query("SELECT IFNULL(SUM(costo_usd),0) c FROM clasificador_ia_costos")[0]["c"]
    except Exception as exc:  # noqa: BLE001
        return _layout("Error", f"<div class=vacio>{_e(exc)}</div>", usuario=usuario)

    def pct(ok, t):
        return f"{(ok or 0) / t * 100:.1f}%" if t else "—"

    proy = float(r["prom"] or 0) * config.filas_mes_estimado
    cards = (
        "<div class=cards>"
        f"<div class=card><div class=n>{r['n']:,}</div><div class=l>Filas comparadas</div></div>"
        f"<div class=card><div class=n>{pct(r['ci'], r['n'])}</div><div class=l>Acierto interés</div></div>"
        f"<div class=card><div class=n>{pct(r['cp'], r['ncp'])}</div><div class=l>Acierto pactivo</div></div>"
        f"<div class=card><div class=n>{pct(r['cc'], r['ncp'])}</div><div class=l>Acierto comp</div></div>"
        f"<div class=card><div class=n>{pct(r['cpr'], r['ncp'])}</div><div class=l>Acierto pres</div></div>"
        f"<div class=card><div class=n>${float(r['costo']):.2f}</div><div class=l>Costo backtest</div></div>"
        f"<div class=card><div class=n>${float(r['prom'] or 0)*1000:.2f}</div><div class=l>$ / 1.000 filas</div></div>"
        f"<div class=card><div class=n>${proy:,.0f}</div><div class=l>Proyección/mes</div></div>"
        "</div>"
    )

    # Cross matrix 2x3 — cards clickeables, color según naturaleza
    matriz_html = ["<h2>Cruz humano vs IA — click en cualquier celda para auditar</h2>",
                   "<div class='matriz'>"]
    for name, (etiq, color, _) in _CELDAS.items():
        n = matriz.get(name, 0) or 0
        matriz_html.append(
            f"<a class='mcell m-{color}' href='/comparacion?cell={name}'>"
            f"<div class=mn>{n:,}</div>"
            f"<div class=ml>{etiq}</div>"
            f"</a>"
        )
    matriz_html.append("</div>")

    # Tabla por método con link
    metodo_rows = []
    for m in por_metodo:
        nm = m["ia_metodo"] or "?"
        etiqueta = _METODOS.get(nm, nm)
        ai = pct(m["ci"], m["n"])
        ap = pct(m["cp"], m["ncp"]) if m["ncp"] else "—"
        costo = float(m["costo"] or 0)
        metodo_rows.append(
            f"<tr><td><a href='/comparacion?metodo={nm}'>{_e(etiqueta)}</a></td>"
            f"<td>{m['n']:,}</td><td>{ai}</td><td>{ap}</td>"
            f"<td>${costo:.4f}</td></tr>"
        )
    metodo_html = (
        "<h2>Por etapa de la cascada</h2><table>"
        "<tr><th>Vía</th><th>Filas</th><th>Acierto interés</th>"
        "<th>Acierto pactivo</th><th>Costo</th></tr>"
        + "".join(metodo_rows) + "</table>"
    )

    # Top pactivos sospechosos — el "Servicio de Aseo, Adjunto, Cocina" de la vida
    if sospechosos:
        sosp_rows = "".join(
            f"<tr><td><a href='/comparacion?pactivo_sug={_e(s['p'])}'>{_e(s['p'])}</a></td>"
            f"<td>{s['n']:,}</td><td>{s['descartes']:,}</td>"
            f"<td><b>{s['pct_desc']}%</b></td></tr>"
            for s in sospechosos
        )
        sosp_html = (
            "<h2>Pactivos sospechosos — IA dijo interés pero humano descarta seguido</h2>"
            "<p style='font-size:13px;color:#6b7689'>Filtro: ≥5 filas con ese pactivo y ≥50% descartadas. "
            "Candidatos a regla automática \"este pactivo → descarte\".</p>"
            "<table><tr><th>Pactivo sugerido por IA</th><th>Filas</th>"
            "<th>Descartes humanos</th><th>% descarte</th></tr>"
            + sosp_rows + "</table>"
        )
    else:
        sosp_html = ("<h2>Pactivos sospechosos</h2><div class=vacio>"
                     "Sin pactivos con ≥50% de descartes humanos (mínimo 5 filas). 🎉</div>")

    coste_acum = (
        f"<p style='font-size:13px;color:#6b7689'>Gasto total acumulado de la cuenta: "
        f"<b>${float(acum):.2f}</b> de ${config.budget_usd:.0f} (mensual)</p>"
    )

    cuerpo = (
        "<h1>Backtest · IA vs personas</h1>"
        + cards + coste_acum
        + "".join(matriz_html)
        + metodo_html
        + sosp_html
    )
    return _layout("Backtest", cuerpo, usuario=usuario)


def _comparacion_listado(cell: str, metodo: str, pactivo_sug: str, limit: int,
                         usuario: dict | None = None) -> str:
    """Drill-down: muestra las filas que caen en un filtro específico."""
    limit = max(20, min(500, limit))
    cond = []
    args: list = []
    titulo_bits = []
    if cell in _CELDAS:
        etiq, _, sql = _CELDAS[cell]
        cond.append(sql)
        titulo_bits.append(etiq)
    if metodo in _METODOS:
        cond.append("ia_metodo=%s")
        args.append(metodo)
        titulo_bits.append(f"vía {_METODOS[metodo]}")
    if pactivo_sug:
        cond.append("ia_pactivo=%s")
        args.append(pactivo_sug)
        titulo_bits.append(f"pactivo IA = «{pactivo_sug}»")

    where = " AND ".join(cond) if cond else "1=1"
    try:
        total = _query(
            f"SELECT COUNT(*) n FROM clasificador_ia_backtest WHERE {where}",
            tuple(args),
        )[0]["n"]
        filas = _query(
            "SELECT tabla_origen, fila_id, descripcion, humano_estado_gestor, "
            "humano_pactivo, humano_composicion, humano_presentacion, "
            "ia_interes, ia_pactivo, ia_composicion, ia_presentacion, "
            "ia_metodo, ia_razon, ia_pactivo_nuevo, creado_en "
            f"FROM clasificador_ia_backtest WHERE {where} "
            "ORDER BY creado_en DESC LIMIT %s",
            tuple(args) + (limit,),
        )
    except Exception as exc:  # noqa: BLE001
        return _layout("Error", f"<div class=vacio>{_e(exc)}</div>", usuario=usuario)

    bloques = []
    for f in filas:
        ia_pact = f.get("ia_pactivo") or f.get("ia_pactivo_nuevo") or "—"
        hp = f.get("humano_pactivo") or "—"
        ia_int = "interés" if f.get("ia_interes") == 1 else "descarte"
        hu_int = "interés" if f.get("humano_estado_gestor") == 1 else "descarte"
        bloques.append(
            f"<div class=fila>"
            f"<div class=meta><b>{_e(f['tabla_origen'])} #{f['fila_id']}</b> · "
            f"<span class='badge b-met'>vía {_e(_METODOS.get(f.get('ia_metodo'), f.get('ia_metodo') or '?'))}</span>"
            f"</div>"
            f"<div class=desc>{_e((f.get('descripcion') or '')[:280])}</div>"
            f"<div class=razon>IA: {_e(f.get('ia_razon') or '')}</div>"
            "<div class=lineh>"
            f"<div><b>Humano</b> ({hu_int}): {_e(hp)}"
            f" · {_e(f.get('humano_composicion') or '—')}"
            f" · {_e(f.get('humano_presentacion') or '—')}</div>"
            f"<div><b>IA</b> ({ia_int}): {_e(ia_pact)}"
            f" · {_e(f.get('ia_composicion') or '—')}"
            f" · {_e(f.get('ia_presentacion') or '—')}</div>"
            "</div></div>"
        )

    titulo = " · ".join(titulo_bits) or "(todas)"
    qs_paginar = f"cell={cell}&metodo={metodo}&pactivo_sug={pactivo_sug}"
    cuerpo = (
        f"<h1>Backtest · {_e(titulo)}</h1>"
        f"<p style='font-size:13px'><b>{total:,}</b> filas encuentran este filtro · "
        f"<a href='/comparacion'>← volver al dashboard</a></p>"
        + "".join(bloques)
        + (f"<p style='font-size:13px;color:#6b7689'>Mostrando primeras {limit}. "
           f"Subir LIMIT: <a href='/comparacion?{qs_paginar}&limit=500'>ver 500</a></p>"
           if total > limit else "")
    )
    return _layout("Backtest detalle", cuerpo, usuario=usuario)


# --------------------------------------------------------------- Revisión ---
# Segregación interés vs descarte: un revisor toma `tipo=descarte` (revisar todas
# las descartas) y otro `tipo=interes` (revisar todas las de interés, con su
# pactivo/composición/presentación). `nuevo` es un subconjunto de interés —
# las que Claude propone con un pactivo fuera del catálogo.
_TIPOS = {
    # interés = de interés Y con pactivo del catálogo. Los de pactivo NUEVO
    # tienen su propia categoría aparte (no se mezclan en "interés").
    # 'posible' (Etapa 2 — 2026-06-05): subconjunto de 'interés' cuya rama es
    # 'modelo_marcas_posible' (confianza media 0.30-0.49). Aparece en amarillo
    # y exige que el revisor complete pact+comp+pres antes de aprobar.
    "interes": ("interes_sugerido=1 AND (pactivo_nuevo IS NULL OR pactivo_nuevo='') "
                "AND metodo<>'modelo_marcas_posible'"),
    "posible": "metodo='modelo_marcas_posible'",
    "descarte": "interes_sugerido=0",
    "nuevo": "pactivo_nuevo IS NOT NULL AND pactivo_nuevo<>''",
}

# Columnas que exporta el legacy gestor_licitaciones para cada tabla.
# Las replicamos para que el equipo encuentre lo mismo que ya conoce.
# Source: gestor_2021/app/Http/Controllers/{CompraAgil,LicitacionesDiarias}Controller.php
_LEGACY_COLS = {
    "compra_agil": [
        # (header_legacy, columna_sql_en_compra_agil)
        # Quitadas (no aplican en compra ágil, son propias de licitaciones):
        #   Duración Contrato, Precio Ponderación, Tiempo del contrato,
        #   Garantía seriedad ofertas, Garantía seriedad contrato.
        ("Demandante", "Demandante"),
        ("Unidad de compra", "Unidad_Compra"),
        ("Región", "Region"),
        ("Comuna", "Comuna"),
        ("Fecha Publicación", "Fecha_Publicacion"),
        ("Fecha Cierre", "Fecha_Cierre"),
        ("Unidades", "Cantidad"),
        ("Medida", "Unidad_Medida"),
        ("Descripción", "Descripcion"),
        ("Producto o Servicio a contratar", "Producto_Servicio"),
        ("Licitación", "Licitacion"),
        ("Item", "Item"),
        ("Descripción PHT", "VINCULOS"),
        ("Rut Cliente", "Rut_Cliente"),
        ("Pactivo", "pactivo"),
        ("Composición", "composicion"),
        ("Presentación", "presentacion"),
        ("Nombre del Contacto", "nombre_contacto"),
        ("Teléfono de contacto", "telefono_contacto"),
        ("Mail de contacto", "mail_contacto"),
        ("Monto Total", "monto_total"),
        ("Estado gestor", "estado_gestor"),
        ("Usuario", "nombre_clasificador"),
    ],
    "Licitaciones_diarias": [
        ("Demandante", "Demandante"),
        ("Unidad de compra", "Unidad_Compra"),
        ("Región", "Region"),
        ("Comuna", "Comuna"),
        ("Fecha Publicación", "Fecha_Publicacion"),
        ("Fecha Cierre", "Fecha_Cierre"),
        ("Fecha inicio pregunta", "Fechainiciopregunta"),
        ("Fecha fin pregunta", "Fechafinalpregunta"),
        ("Cod Onu", "Cod_Onu"),
        ("Unidades", "Cantidad"),
        ("Medida", "Unidad_Medida"),
        ("Descripción", "Descripcion"),
        ("Producto o Servicio a contratar", "Producto_Servicio"),
        ("Licitación", "Licitacion"),
        ("Item", "Item"),
        ("Descripción PHT", "VINCULOS"),
        ("Rut Cliente", "Rut_Cliente"),
        ("Duración Contrato", "Duracion_Contrato"),
        ("Precio Ponderación", "Precio_Ponderacion"),
        ("Tiempo del contrato", "Tiempo_Contrato"),
        ("Garantía seriedad ofertas", "Garantia_Seriedad_Ofertas"),
        ("Garantía seriedad contrato", "Garantia_Seriedad_Contrato"),
        ("Fecha Adjudicación", "Fechaadjudicacion"),
        ("Pactivo", "pactivo"),
        ("Composición", "composicion"),
        ("Presentación", "presentacion"),
        ("Estado gestor", "estado_gestor"),
        ("Usuario", "nombre_clasificador"),
    ],
}

# Columnas EXTRA con info de la IA que añadimos al final del legacy export
# para que el revisor vea de un vistazo lo que la cascada propuso.
_IA_EXTRAS_COLS = [
    "IA · Tipo", "IA · Pactivo sugerido", "IA · Composición", "IA · Presentación",
    "IA · Vía (etapa)", "IA · Confianza", "IA · Razón", "IA · Pactivo Nuevo",
    "IA · Revisada", "IA · Revisor", "IA · Fecha revisión", "IA · Acierto humano",
]


# Anchos de columna sugeridos para el XLSX (en unidades de Excel — ~ ancho
# en caracteres). Replican el legacy y dan espacio a campos largos como
# Descripción y Razón. Si el header no figura, se usa el default (15).
_ANCHO_XLSX = {
    "Demandante": 38, "Unidad de compra": 30, "Región": 14, "Comuna": 16,
    "Fecha Publicación": 18, "Fecha Cierre": 18, "Unidades": 10, "Medida": 12,
    "Descripción": 60, "Producto o Servicio a contratar": 35,
    "Licitación": 22, "Item": 16, "Descripción PHT": 50, "Rut Cliente": 14,
    "Duración Contrato": 16, "Precio Ponderación": 15, "Tiempo del contrato": 18,
    "Garantía seriedad ofertas": 18, "Garantía seriedad contrato": 18,
    "Pactivo": 26, "Composición": 18, "Presentación": 18,
    "Nombre del Contacto": 22, "Teléfono de contacto": 16, "Mail de contacto": 28,
    "Monto Total": 14, "Estado gestor": 12, "Usuario": 22,
    "Fecha inicio pregunta": 18, "Fecha fin pregunta": 18, "Cod Onu": 14,
    "Fecha Adjudicación": 18,
    "IA · Tipo": 14, "IA · Pactivo sugerido": 26, "IA · Composición": 16,
    "IA · Presentación": 18, "IA · Vía (etapa)": 22, "IA · Confianza": 12,
    "IA · Razón": 55, "IA · Pactivo Nuevo": 22, "IA · Revisada": 10,
    "IA · Revisor": 22, "IA · Fecha revisión": 18, "IA · Acierto humano": 14,
}


@app.get("/revision.xlsx")
def revision_xlsx(tabla: str = "", tipo: str = "", metodo: str = "", conf: str = "",
                  rango: str = "ayer_hoy", desde: str = "", hasta: str = "",
                  estado: str = "pendientes", busqueda: str = "",
                  licitacion: str = ""):
    """Export ESTILO LEGACY como XLSX con formato — replica el Excel del
    gestor_licitaciones (CompraAgilController->generarExcelCa /
    LicitacionesDiariasController->generarExcelClasificacion).

    Hereda: 28 columnas del legacy + 12 columnas IA al final.
    Agrega: header azul corporativo, autofilter, freeze del row 1, anchos
    fijos por columna, wrap text para descripciones, pintado por estado
    (verde=interés, rojo=descarte, gris=pendiente) — mismo que el legacy.

    Si la tabla no se filtra, default a compra_agil (más común)."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    estado = estado if estado in _ESTADOS else "pendientes"
    if rango == "todas":
        rango = ""
    # Si no hay filtro de tabla, default a compra_agil para usar su schema legacy
    tabla_export = tabla if tabla in TABLAS_VALIDAS else "compra_agil"
    cols_def = _LEGACY_COLS[tabla_export]

    # WHERE sobre el log
    cond = [_ESTADOS[estado], "log.tabla_origen = %s"]
    args: list = [tabla_export]
    if tipo in _TIPOS:
        cond.append(_TIPOS[tipo].replace("interes_sugerido", "log.interes_sugerido")
                                .replace("pactivo_nuevo", "log.pactivo_nuevo"))
    if metodo in _METODOS:
        cond.append("log.metodo=%s"); args.append(metodo)
    if conf == "baja":
        cond.append("log.confianza < 0.7")
    elif conf == "media":
        cond.append("log.confianza >= 0.7 AND log.confianza < 0.85")
    elif conf == "alta":
        cond.append("log.confianza >= 0.85")
    if rango in _RANGOS and not desde and not hasta:
        cond.append(_RANGOS[rango].replace("creado_en", "log.creado_en"))
    _d = _norm_fecha(desde, fin=False)
    _h = _norm_fecha(hasta, fin=True)
    if _d:
        cond.append("log.creado_en >= %s"); args.append(_d)
    if _h:
        cond.append("log.creado_en <= %s"); args.append(_h)
    if busqueda:
        cond.append("log.descripcion LIKE %s"); args.append(f"%{busqueda.strip()}%")
    if licitacion:
        ca_ids, li_ids = _fila_ids_por_licitacion(licitacion.strip())
        ids = ca_ids if tabla_export == "compra_agil" else li_ids
        if not ids:
            cond.append("1=0")
        else:
            ph = ",".join(["%s"] * len(ids))
            cond.append(f"log.fila_id IN ({ph})")
            args.extend(ids)
    where = " AND ".join(cond)

    # SQL: JOIN log con tabla origen para traer las columnas legacy
    col_list = ", ".join(f"t.`{sql_col}` AS `{sql_col}`" for _, sql_col in cols_def)
    sql = (
        f"SELECT {col_list}, "
        "log.interes_sugerido AS ia_int, log.pactivo_sugerido AS ia_pact, "
        "log.composicion_sugerida AS ia_comp, log.presentacion_sugerida AS ia_pres, "
        "log.metodo AS ia_metodo, log.confianza AS ia_conf, log.razon AS ia_razon, "
        "log.pactivo_nuevo AS ia_pactivo_nuevo, log.revisado AS ia_revisado, "
        "log.revisado_por AS ia_revisor, log.revisado_en AS ia_revisado_en, "
        "log.feedback_correcto AS ia_correcto "
        f"FROM clasificador_ia_log log JOIN `{tabla_export}` t ON t.id = log.fila_id "
        f"WHERE {where} ORDER BY log.creado_en DESC LIMIT 50000"
    )
    filas = _query(sql, tuple(args))

    # ----- Construir XLSX con openpyxl -----
    wb = Workbook()
    ws = wb.active
    ws.title = tabla_export[:30]  # límite XLSX

    headers = [h for h, _ in cols_def] + _IA_EXTRAS_COLS

    # Estilos
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="4281C2")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_interes = PatternFill("solid", fgColor="BBDFB9")
    fill_descarte = PatternFill("solid", fgColor="FFBABA")
    fill_pendiente = PatternFill("solid", fgColor="E9E9E9")
    cell_align = Alignment(vertical="top", wrap_text=True)

    # Header
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
    ws.row_dimensions[1].height = 32

    # Anchos de columna
    for col_idx, h in enumerate(headers, 1):
        w = _ANCHO_XLSX.get(h, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    def _fmt(v):
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d %H:%M")
        return v

    def _estado_legible(v):
        return {1: "interés", 0: "descarte"}.get(v, "pendiente")

    # Filas — replicando el coloreo del legacy por estado_gestor
    for row_idx, f in enumerate(filas, start=2):
        row_values = [_fmt(f.get(sql_col)) for _, sql_col in cols_def]
        # Estado gestor legible
        for i, (_, sql_col) in enumerate(cols_def):
            if sql_col == "estado_gestor":
                row_values[i] = _estado_legible(f.get("estado_gestor"))
        # IA extras
        tipo_ia = "INTERÉS" if f.get("ia_int") == 1 else "descarte"
        if (f.get("ia_pactivo_nuevo") or "").strip():
            tipo_ia = "PACTIVO NUEVO"
        revisada = bool(f.get("ia_revisado"))
        row_values += [
            tipo_ia,
            _fmt(f.get("ia_pact")),
            _fmt(f.get("ia_comp")),
            _fmt(f.get("ia_pres")),
            _METODOS.get(f.get("ia_metodo"), f.get("ia_metodo") or ""),
            round(float(f.get("ia_conf") or 0), 2),
            _fmt(f.get("ia_razon")),
            _fmt(f.get("ia_pactivo_nuevo")),
            "sí" if revisada else "no",
            _fmt(f.get("ia_revisor")),
            _fmt(f.get("ia_revisado_en")),
            ("sí" if f.get("ia_correcto") == 1 else "no") if revisada else "",
        ]
        # Pintar la fila según estado_gestor (igual que el legacy)
        est = f.get("estado_gestor")
        row_fill = (fill_interes if est == 1 else
                    fill_descarte if est == 0 else fill_pendiente)
        for col_idx, v in enumerate(row_values, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.fill = row_fill
            c.alignment = cell_align

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    if filas:
        ws.auto_filter.ref = ws.dimensions

    # Guardar a BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"{tabla_export}-{datetime.now():%Y%m%d-%H%M}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


def qs_extras_str(tabla, tipo, metodo, conf, rango, desde, hasta, estado, busqueda, licitacion, tipo_rev=""):
    """Querystring común para preservar filtros al armar links/export."""
    pares = [
        ("tabla", tabla), ("tipo", tipo), ("metodo", metodo), ("conf", conf),
        ("rango", rango), ("desde", desde), ("hasta", hasta),
        ("estado", estado if estado != "pendientes" else ""),
        ("busqueda", busqueda), ("licitacion", licitacion),
        ("tipo_rev", tipo_rev),
    ]
    return "&".join(f"{k}={v}" for k, v in pares if v)


# Etiquetas legibles de cada etapa de la cascada que resolvió la fila.
_METODOS = {
    "cruce_base": "cruce Base",
    "descarte_item": "descarte por rubro",
    "modelo_descarte": "clasif. de descarte",
    "modelo_pactivo": "clasif. de pactivo",
    "conflicto_regla_modelo": "regla vs. modelo",
    "historico": "histórico",
    "regla_diccionario": "reglas",
    "claude": "Claude",
}


_ESTADOS = {"pendientes": "revisado=0", "revisadas": "revisado=1", "todas": "1=1"}
# Subfiltros dentro de "revisadas" — para contar cuántas se aprobaron, se
# corrigieron (cambiaron el pactivo) o se descartaron en un rango de fechas.
_TIPOS_REV = {
    "aprobadas":   "feedback_correcto=1",
    # 'corregidas' = TODAS las que el humano cambió algo respecto a la IA:
    #   - cambió el pactivo (feedback_pactivo lleno)
    #   - cambió la decisión a descarte (feedback_pactivo vacío)
    #   - rescató un descarte poniendo pactivo
    # Pedido del usuario 2026-06-02: que aparezcan ambas en el mismo filtro.
    "corregidas":  "feedback_correcto=0",
    "descartadas": "(feedback_correcto=0 AND (feedback_pactivo IS NULL "
                   "OR TRIM(feedback_pactivo)=''))",
}
_RANGOS = {
    "ayer_hoy": "creado_en >= CURDATE() - INTERVAL 1 DAY",
    "hoy": "creado_en >= CURDATE()",
    "ayer": "creado_en >= CURDATE() - INTERVAL 1 DAY AND creado_en < CURDATE()",
    "semana": "creado_en >= CURDATE() - INTERVAL 7 DAY",
    "mes": "creado_en >= CURDATE() - INTERVAL 30 DAY",
}


# Supergrupos para ordenar visualmente la cola. El revisor ve primero los
# clusters claros (interés con histórico/cruce que son los más confiables) y
# después los que requieren más atención (Claude, conflictos). Adjuntos y
# pactivos nuevos van como categorías propias.
def _supergrupo(fila: dict) -> tuple:
    """Devuelve (clave_orden, etiqueta_visible) para agrupar la fila."""
    interes = fila.get("interes_sugerido")
    metodo = fila.get("metodo") or ""
    tabla = fila.get("tabla_origen") or ""
    pact = (fila.get("pactivo_sugerido") or "").strip()
    es_nuevo = bool((fila.get("pactivo_nuevo") or "").strip())

    if es_nuevo:
        return ("z1_nuevo", "⚠ PACTIVOS NUEVOS · Claude propuso fuera de catálogo")
    if pact == "Adjunto" and tabla == "compra_agil":
        return ("y1_adj_ca", "📎 ADJUNTOS · COMPRAS ÁGILES")
    if pact == "Adjunto" and tabla == "Licitaciones_diarias":
        return ("y2_adj_li", "📎 ADJUNTOS · LICITACIONES")
    if interes == 1:
        if metodo in ("cruce_base", "historico"):
            return ("a1_int_hist", "🟢 INTERÉS · Cruce histórico (OC reales + descripción ya clasificada)")
        if metodo in ("modelo_pactivo",):
            return ("a2_int_ml", "🟢 INTERÉS · Clasificación ML (modelo entrenado de pactivo)")
        if metodo == "regla_diccionario":
            return ("a3_int_reglas", "🟢 INTERÉS · Reglas (match diccionario)")
        if metodo == "claude":
            return ("a4_int_claude", "🟢 INTERÉS · Claude")
        return ("a9_int_otro", "🟢 INTERÉS · otro")
    if interes == 0:
        if metodo == "descarte_item":
            return ("b1_desc_item", "🔴 DESCARTE · Por rubro (códigos siempre descartados)")
        if metodo == "modelo_descarte":
            return ("b2_desc_ml", "🔴 DESCARTE · Clasificador entrenado")
        if metodo == "conflicto_regla_modelo":
            return ("b3_desc_conf", "🔴 DESCARTE · Conflicto regla vs modelo")
        if metodo == "historico":
            return ("b4_desc_hist", "🔴 DESCARTE · Histórico humano")
        if metodo == "claude":
            return ("b5_desc_claude", "🔴 DESCARTE · Claude")
        return ("b9_desc_otro", "🔴 DESCARTE · otro")
    return ("zz", "otro")


def _fila_ids_por_licitacion(numero: str) -> "tuple[list, list]":
    """Busca el número de licitación/compra ágil en las 2 tablas origen y
    devuelve (compra_agil_ids, licitaciones_ids). El revisor escribe el número
    en el filtro y vemos exactamente esa fila. Match LIKE para tolerar prefijos."""
    if not numero:
        return ([], [])
    out = ([], [])
    for i, t in enumerate(TABLAS_VALIDAS):
        try:
            r = _query(
                f"SELECT id FROM `{t}` WHERE Licitacion LIKE %s LIMIT 200",
                (f"%{numero}%",),
            )
            out[i].extend(x["id"] for x in r if "id" in x)
        except Exception:  # noqa: BLE001
            pass
    return out


@app.get("/revision", response_class=HTMLResponse)
def revision(request: Request, hoja: int = 1, msg: str = "", tabla: str = "",
             tipo: str = "", metodo: str = "", conf: str = "", por_hoja: int = 0,
             estado: str = "pendientes", rango: str = "",
             desde: str = "", hasta: str = "",
             busqueda: str = "", licitacion: str = "",
             tipo_rev: str = "", rev_desde: str = "", rev_hasta: str = "") -> str:
    usuario = usuario_actual(request)
    es_admin = _es_admin(usuario)
    hoja = max(1, hoja)
    por_hoja = por_hoja if por_hoja in POR_HOJA_OPCIONES else POR_HOJA_DEFAULT
    estado = estado if estado in _ESTADOS else "pendientes"
    # Sin preset: la fecha se controla con desde/hasta. Si el revisor entra a
    # /revision sin params, aplicamos el default ANTES de construir el WHERE
    # (no solo en el input visual) — así el conteo y la cola coinciden con lo
    # que muestra el input desde el primer load (sin esperar a "aplicar").
    rango = ""
    from datetime import date as _date_d, timedelta as _td_d
    # Cuando el enlace trae un filtro por FECHA DE REVISIÓN (rev_desde/rev_hasta,
    # p.ej. desde el reporte diario → "ver los errores del día"), NO forzamos el
    # rango default de fecha de publicación: si no, ambos filtros se cruzarían y
    # se perderían errores cuya publicación es anterior a ayer.
    _hay_rev_filtro = bool(rev_desde or rev_hasta)
    if not desde and not _hay_rev_filtro:
        desde = f"{(_date_d.today() - _td_d(days=1)).isoformat()}T18:30"
    if not hasta and not _hay_rev_filtro:
        hasta = f"{_date_d.today().isoformat()}T18:30"
    cond = [_ESTADOS[estado]]
    args: list = []
    if tabla in TABLAS_VALIDAS:
        cond.append("tabla_origen=%s")
        args.append(tabla)
    if tipo in _TIPOS:
        cond.append(_TIPOS[tipo])
    if metodo in _METODOS:
        cond.append("metodo=%s")
        args.append(metodo)
    # Subfiltro dentro de "revisadas": aprobadas / corregidas / descartadas.
    if estado == "revisadas" and tipo_rev in _TIPOS_REV:
        cond.append(_TIPOS_REV[tipo_rev])
    # Filtro por banda de confianza — para priorizar los casos dudosos (<0.7)
    # o, al revés, repasar masivamente los seguros (>=0.85) con un solo OK.
    if conf == "baja":
        cond.append("confianza < 0.7")
    elif conf == "media":
        cond.append("confianza >= 0.7 AND confianza < 0.85")
    elif conf == "alta":
        cond.append("confianza >= 0.85")
    # Fecha — por FECHA DE PUBLICACIÓN (cuando salió en mercadopublico), NO por
    # creado_en (cuándo la IA la procesó). La fecha de publicación vive en la
    # tabla origen — subconsulta correlacionada por PK. Acepta fecha sola o
    # datetime-local para cortes por hora/turno.
    _SQL_PUB_FECHA = (
        "COALESCE("
        "(SELECT ca.Fecha_Publicacion FROM compra_agil ca "
        "WHERE ca.id=clasificador_ia_log.fila_id "
        "AND clasificador_ia_log.tabla_origen='compra_agil'),"
        "(SELECT ld.Fecha_Publicacion FROM Licitaciones_diarias ld "
        "WHERE ld.id=clasificador_ia_log.fila_id "
        "AND clasificador_ia_log.tabla_origen='Licitaciones_diarias'))"
    )
    _d = _norm_fecha(desde, fin=False)
    _h = _norm_fecha(hasta, fin=True)
    if _d:
        cond.append(f"{_SQL_PUB_FECHA} >= %s")
        args.append(_d)
    if _h:
        cond.append(f"{_SQL_PUB_FECHA} <= %s")
        args.append(_h)
    # Fecha de REVISIÓN (revisado_en) — para enlaces del reporte diario que
    # llevan exactamente a los casos que el humano revisó/corrigió ese día.
    _rd = _norm_fecha(rev_desde, fin=False)
    _rh = _norm_fecha(rev_hasta, fin=True)
    if _rd:
        cond.append("revisado_en >= %s")
        args.append(_rd)
    if _rh:
        cond.append("revisado_en <= %s")
        args.append(_rh)
    # Búsqueda de texto en la descripción (LIKE %X%)
    if busqueda:
        cond.append("descripcion LIKE %s")
        args.append(f"%{busqueda.strip()}%")
    # Filtro por número de licitación / compra ágil: pre-resuelve los fila_id
    # en la tabla origen para no JOIN-ear en cada query.
    if licitacion:
        ca_ids, li_ids = _fila_ids_por_licitacion(licitacion.strip())
        partes = []
        if ca_ids:
            ph = ",".join(["%s"] * len(ca_ids))
            partes.append(f"(tabla_origen='compra_agil' AND fila_id IN ({ph}))")
            args.extend(ca_ids)
        if li_ids:
            ph = ",".join(["%s"] * len(li_ids))
            partes.append(f"(tabla_origen='Licitaciones_diarias' AND fila_id IN ({ph}))")
            args.extend(li_ids)
        if not partes:
            cond.append("1=0")  # nada encontrado → no devuelve nada
        else:
            cond.append("(" + " OR ".join(partes) + ")")
    # COLA SINCRONIZADA CON EL LEGACY: una fila que un HUMANO ya clasificó o
    # descartó en gestor_licitaciones (estado_gestor seteado + firma humana) ya
    # está resuelta — NO debe aparecer pendiente acá, aunque la IA la procesó.
    # Los dos sistemas corren en paralelo. Es un EXISTS correlacionado por PK (id)
    # contra la MISMA BD del clásico — NO hay llamada a otro sistema; es un index
    # lookup, no se "queda pegado". El worker IA igual procesa TODO (esté o no
    # clasificado en el legacy); esto solo filtra lo que el revisor VE como pendiente.
    if estado == "pendientes":
        cond.append(
            "NOT EXISTS (SELECT 1 FROM compra_agil ca "
            "WHERE ca.id=clasificador_ia_log.fila_id "
            "AND clasificador_ia_log.tabla_origen='compra_agil' "
            "AND ca.estado_gestor IS NOT NULL AND ca.nombre_clasificador IS NOT NULL "
            "AND ca.nombre_clasificador NOT REGEXP '^(Bot|BOT|IA_|Bot Eliminado)')"
        )
        cond.append(
            "NOT EXISTS (SELECT 1 FROM Licitaciones_diarias ld "
            "WHERE ld.id=clasificador_ia_log.fila_id "
            "AND clasificador_ia_log.tabla_origen='Licitaciones_diarias' "
            "AND ld.estado_gestor IS NOT NULL AND ld.nombre_clasificador IS NOT NULL "
            "AND ld.nombre_clasificador NOT REGEXP '^(Bot|BOT|IA_|Bot Eliminado)')"
        )
    where = " AND ".join(cond)
    # Para REVISADAS ordeno por revisado_en DESC (lo más recientemente cerrado
    # primero — es lo que el revisor quiere ver para auditar). Para pendientes,
    # ordeno PRIMERO por supergrupo (interés cruce histórico va primero,
    # claude/conflictos al final) — el revisor procesa por bloques afines.
    # Subconsulta correlacionada: la fecha de cierre vive en la tabla origen.
    _SQL_CIERRE = (
        "COALESCE("
        "(SELECT ca.Fecha_Cierre FROM compra_agil ca WHERE ca.id=clasificador_ia_log.fila_id "
        "AND clasificador_ia_log.tabla_origen='compra_agil'),"
        "(SELECT ld.Fecha_Cierre FROM Licitaciones_diarias ld WHERE ld.id=clasificador_ia_log.fila_id "
        "AND clasificador_ia_log.tabla_origen='Licitaciones_diarias'))"
    )
    if estado == "revisadas":
        orden = "revisado_en DESC"
    elif not es_admin:
        # APROBADORES: orden GLOBAL por fecha de cierre más próxima (urgente
        # clasificar lo que cierra antes). Sin fecha → al final. La cola ya viene
        # acotada por el filtro legacy, así que el costo del ORDER es bajo.
        orden = f"{_SQL_CIERRE} IS NULL, {_SQL_CIERRE} ASC"
    else:
        # Orden compuesto por supergrupo (lógico en SQL: CASE WHEN) — evita
        # re-ordenar en Python después. El orden por confianza adentro de cada
        # grupo sigue siendo válido (lo dudoso primero).
        orden = """
        CASE
          WHEN pactivo_nuevo IS NOT NULL AND pactivo_nuevo<>'' THEN 'z1'
          WHEN pactivo_sugerido='Adjunto' AND tabla_origen='compra_agil' THEN 'y1'
          WHEN pactivo_sugerido='Adjunto' AND tabla_origen='Licitaciones_diarias' THEN 'y2'
          WHEN interes_sugerido=1 AND metodo IN ('cruce_base','historico') THEN 'a1'
          WHEN interes_sugerido=1 AND metodo='modelo_pactivo' THEN 'a2'
          WHEN interes_sugerido=1 AND metodo='regla_diccionario' THEN 'a3'
          WHEN interes_sugerido=1 AND metodo='claude' THEN 'a4'
          WHEN interes_sugerido=0 AND metodo='descarte_item' THEN 'b1'
          WHEN interes_sugerido=0 AND metodo='modelo_descarte' THEN 'b2'
          WHEN interes_sugerido=0 AND metodo='conflicto_regla_modelo' THEN 'b3'
          WHEN interes_sugerido=0 AND metodo='historico' THEN 'b4'
          WHEN interes_sugerido=0 AND metodo='claude' THEN 'b5'
          ELSE 'zz'
        END, confianza ASC, creado_en DESC
        """
    try:
        total = _query(
            f"SELECT COUNT(*) n FROM clasificador_ia_log WHERE {where}", tuple(args)
        )[0]["n"]
        filas = _query(
            "SELECT id, tabla_origen, fila_id, descripcion, interes_sugerido, "
            "pactivo_sugerido, composicion_sugerida, presentacion_sugerida, "
            "confianza, razon, pactivo_nuevo, metodo, creado_en, revisado, "
            "revisado_por, revisado_en, feedback_correcto, feedback_pactivo, "
            "feedback_notas FROM clasificador_ia_log "
            f"WHERE {where} ORDER BY {orden} LIMIT %s OFFSET %s",
            tuple(args) + (por_hoja, (hoja - 1) * por_hoja),
        )
    except Exception as exc:  # noqa: BLE001
        return _layout("Error", f"<div class=vacio>{_e(exc)}</div>", usuario=usuario)

    # número de licitación / compra ágil de cada fila — vive en la tabla origen,
    # no en el log; se trae con una consulta por tabla (no una por fila).
    num_lic: dict = {}
    info_origen: dict = {}  # (tabla, id) -> {cierre, vinculos} para la vista de aprobación
    por_tabla: dict = {}
    for f in filas:
        por_tabla.setdefault(f["tabla_origen"], []).append(f["fila_id"])
    for t, ids in por_tabla.items():
        if t in TABLAS_VALIDAS and ids:
            ph = ",".join(["%s"] * len(ids))
            try:
                for r in _query(
                    f"SELECT id, Licitacion, Fecha_Cierre, Fecha_Publicacion, "
                    f"Titulo, Demandante FROM `{t}` WHERE id IN ({ph})", tuple(ids)
                ):
                    num_lic[(t, r["id"])] = r["Licitacion"]
                    info_origen[(t, r["id"])] = {
                        "cierre": r.get("Fecha_Cierre"),
                        "publicacion": r.get("Fecha_Publicacion"),
                        "titulo": r.get("Titulo"),
                        "demandante": r.get("Demandante"),
                    }
            except Exception:  # noqa: BLE001
                pass

    # barra de filtros (mantiene el estado de los otros filtros en cada enlace).
    # "estado" no usa default-vacío como los otros — siempre tiene valor.
    def filtro_link(clave: str, valor: str, etiqueta: str, activo: bool) -> str:
        cur = {"tabla": tabla, "tipo": tipo, "metodo": metodo,
               "conf": conf, "rango": rango, "desde": desde, "hasta": hasta,
               "busqueda": busqueda, "licitacion": licitacion,
               "tipo_rev": tipo_rev, "rev_desde": rev_desde, "rev_hasta": rev_hasta,
               "estado": estado if estado != "pendientes" else "",
               "por_hoja": str(por_hoja) if por_hoja != POR_HOJA_DEFAULT else ""}
        cur[clave] = valor
        qs = "&".join(f"{k}={v}" for k, v in cur.items() if v)
        cls = " class=on" if activo else ""
        return f"<a href='/revision?{qs}'{cls}>{etiqueta}</a>"

    # Filtros: una sola <form> con <select>+inputs. Cada select submitea on
    # change. Mucho más limpio que la pared de links anterior. Los selects
    # tienen typeahead nativo del navegador — si la lista es larga (Vía: 8
    # opciones), apretar la primera letra salta a la opción.
    def opt(value: str, label: str, current) -> str:
        sel = " selected" if str(value) == str(current) else ""
        return f'<option value="{_e(value)}"{sel}>{_e(label)}</option>'

    csv_qs = qs_extras_str(tabla, tipo, metodo, conf, rango, desde, hasta,
                            estado, busqueda, licitacion, tipo_rev)

    # Prellenado de los campos de fecha: si el revisor no fijó un rango propio,
    # mostramos el rango efectivo que se está cargando (ayer 00:00 → hoy 23:59)
    # en formato datetime-local, para que VEA qué está revisando y pueda ajustar
    # el corte por fecha/hora (turnos). Editable: si cambia los campos, mandan ellos.
    from datetime import date as _date, timedelta as _td
    def _dtlocal(v):
        return (v or "").strip().replace(" ", "T")[:16]
    # Default: ventana de cierre de jornada — ayer 18:30 → hoy 18:30 (ventana de
    # 24h con corte a las 18:30, el momento de cierre que el equipo usa). El
    # revisor puede ajustar a NOW si quiere ver lo más reciente.
    if desde:
        desde_val = _dtlocal(desde)
    else:
        desde_val = f"{(_date.today() - _td(days=1)).isoformat()}T18:30"
    if hasta:
        hasta_val = _dtlocal(hasta)
    else:
        hasta_val = f"{_date.today().isoformat()}T18:30"

    # Subfiltro VEREDICTO solo cuando estado=revisadas (aprobadas/corregidas/desc).
    sel_veredicto = ""
    if estado == "revisadas":
        sel_veredicto = (
            "<label>Veredicto</label>"
            "<select name=tipo_rev onchange='this.form.submit()'>"
            + opt("", "todas", tipo_rev)
            + opt("aprobadas", "✓ aprobadas", tipo_rev)
            + opt("corregidas", "✏ corregidas", tipo_rev)
            + opt("descartadas", "✗ descartadas", tipo_rev)
            + "</select>"
        )

    filtros = (
        f"<form method=get action='/revision' class=ff>"
        # ---- línea 1: Estado · Tabla · Tipo
        "<div class=fila-filt>"
        "<label>Estado</label>"
        f"<select name=estado onchange='this.form.submit()'>"
        + opt("pendientes", "pendientes", estado)
        + opt("revisadas", "revisadas", estado)
        + opt("todas", "todas", estado)
        + "</select>"
        + sel_veredicto
        + "<label>Tabla</label>"
        f"<select name=tabla onchange='this.form.submit()'>"
        + opt("", "todas", tabla)
        + opt("compra_agil", "compra ágil", tabla)
        + opt("Licitaciones_diarias", "licitaciones", tabla)
        + "</select>"
        "<label>Tipo</label>"
        f"<select name=tipo onchange='this.form.submit()'>"
        + opt("", "todos", tipo)
        + opt("interes", "interés", tipo)
        + opt("posible", "posiblemente de interés", tipo)
        + opt("descarte", "descarte", tipo)
        + opt("nuevo", "pactivo nuevo", tipo)
        + "</select>"
        "</div>"
        # ---- línea 2: Fecha de PUBLICACIÓN (corte por hora para turnos)
        "<div class=fila-filt>"
        "<label>Publicada desde</label>"
        f"<input type=datetime-local name=desde value='{_e(desde_val)}'>"
        f"<label>hasta</label><input type=datetime-local name=hasta value='{_e(hasta_val)}'>"
        "<button type=submit class=sec>aplicar</button>"
        "</div>"
        # ---- línea 3: Buscar (texto) · N° licitación · Excel
        "<div class=fila-filt>"
        "<label>Buscar</label>"
        f"<input type=text name=busqueda value='{_e(busqueda)}' "
        f"placeholder='palabra en la glosa…' style='width:240px'>"
        "<label>N° licit/compra</label>"
        f"<input type=text name=licitacion value='{_e(licitacion)}' "
        f"placeholder='ej. 5523-145-L226' style='width:200px'>"
        "<button type=submit class=sec>buscar</button>"
        f"&nbsp;<a class=btn-excel href='/revision.xlsx?{_e(csv_qs)}'>📥 Excel (XLSX)</a>"
        "</div>"
        # ---- línea 4: Confianza · Vía · Por hoja
        "<div class=fila-filt>"
        "<label>Confianza</label>"
        f"<select name=conf onchange='this.form.submit()'>"
        + opt("", "toda", conf)
        + opt("baja", "< 0.70 (duda)", conf)
        + opt("media", "0.70 - 0.85", conf)
        + opt("alta", "≥ 0.85 (segura)", conf)
        + "</select>"
        "<label>Vía</label>"
        f"<select name=metodo onchange='this.form.submit()'>"
        + opt("", "todas", metodo)
        + "".join(opt(k, v, metodo) for k, v in _METODOS.items())
        + "</select>"
        "<label>Por hoja</label>"
        f"<select name=por_hoja onchange='this.form.submit()'>"
        + "".join(opt(str(n), str(n), str(por_hoja)) for n in POR_HOJA_OPCIONES)
        + "</select>"
        "</div>"
        "</form>"
    )

    def _fmt_dt(dt):
        # Defensivo: Fecha_Cierre/Fecha_Publicacion a veces vuelven como string
        # desde la BD (columna VARCHAR mal tipada). Si no es datetime, mostrar
        # el string tal cual — antes esto crasheaba el panel (AttributeError).
        if not dt:
            return "—"
        if isinstance(dt, str):
            return dt[:16] if dt else "—"
        try:
            return dt.strftime("%d-%m-%Y %H:%M")
        except Exception:  # noqa: BLE001
            return str(dt)[:16]

    # Estado del worker: HASTA QUÉ PUBLICACIÓN está al día (frontera) + última
    # fila procesada + cuántas le faltan. La FRONTERA es lo más importante: dice
    # "todo lo publicado antes de esta fecha YA está procesado". Si la frontera
    # está cerca de NOW (≤15min) el sistema está al día; si está muy atrás, hay
    # delay y NO podemos confiar en "vi todo lo publicado hasta las 18:30".
    estado_worker = ""
    try:
        from datetime import datetime as _dt, timedelta as _tdh
        # El container corre en UTC pero el usuario está en Chile (UTC-4). Los
        # timestamps del LOG (creado_en, escritos con datetime.now() Python) están
        # en UTC; Fecha_Publicacion viene del scraping y ya está en hora Chile.
        _TZ_OFFSET = _tdh(hours=4)  # UTC -> Chile
        _now_chile = _dt.now() - _TZ_OFFSET
        # ult fila procesada (creado_en está en UTC → convertir a Chile)
        _ult_utc = _query("SELECT MAX(creado_en) u FROM clasificador_ia_log")[0]["u"]
        _ult = (_ult_utc - _TZ_OFFSET) if _ult_utc else None
        _now = _now_chile
        # frontera = MIN(Fecha_Publicacion) de las pendientes (estado_gestor NULL
        # + NO en log). Todo lo publicado ANTES está procesado.
        # Filtro `> '2000-01-01'`: ignora las zero-dates ('0000-00-00 00:00:00')
        # que pymysql devolvería como string y romperían .total_seconds(). Hay
        # 232 zero-dates en Licitaciones_diarias.Fecha_Cierre y la prevención
        # cubre también Fecha_Publicacion por robustez.
        _mn_ca = _query(
            "SELECT MIN(t.Fecha_Publicacion) m FROM compra_agil t WHERE t.estado_gestor IS NULL "
            "AND t.Fecha_Publicacion > '2000-01-01' "
            "AND NOT EXISTS (SELECT 1 FROM clasificador_ia_log l "
            "WHERE l.tabla_origen='compra_agil' AND l.fila_id=t.id)"
        )[0]["m"]
        _mn_li = _query(
            "SELECT MIN(t.Fecha_Publicacion) m FROM Licitaciones_diarias t WHERE t.estado_gestor IS NULL "
            "AND t.Fecha_Publicacion > '2000-01-01' "
            "AND NOT EXISTS (SELECT 1 FROM clasificador_ia_log l "
            "WHERE l.tabla_origen='Licitaciones_diarias' AND l.fila_id=t.id)"
        )[0]["m"]
        # la pendiente más antigua entre ambas tablas (descarta cualquier string
        # residual por si MIN devolvió zero-date pese al filtro — defensa final)
        _pend = [x for x in (_mn_ca, _mn_li) if x and not isinstance(x, str)]
        _frontera = min(_pend) if _pend else _now  # si no hay pendientes, al día hasta ahora
        _seg_front = int((_now - _frontera).total_seconds())
        _cls = "ok" if _seg_front < 900 else ("warn" if _seg_front < 3600 else "bad")
        _msg_ok = "AL DÍA" if _seg_front < 900 else ("LEVE DELAY" if _seg_front < 3600 else "DELAY")
        # contadores y "última procesada"
        _falt_ca = _query(
            "SELECT COUNT(*) n FROM compra_agil t WHERE t.estado_gestor IS NULL "
            "AND NOT EXISTS (SELECT 1 FROM clasificador_ia_log l "
            "WHERE l.tabla_origen='compra_agil' AND l.fila_id=t.id)"
        )[0]["n"]
        _falt_li = _query(
            "SELECT COUNT(*) n FROM Licitaciones_diarias t WHERE t.estado_gestor IS NULL "
            "AND NOT EXISTS (SELECT 1 FROM clasificador_ia_log l "
            "WHERE l.tabla_origen='Licitaciones_diarias' AND l.fila_id=t.id)"
        )[0]["n"]
        _ult_txt = f"última fila procesada: {_fmt_dt(_ult)}" if _ult else ""
        estado_worker = (
            f"<div class='worker-st worker-{_cls}'>"
            f"<b>Worker {_msg_ok}</b> · todo lo publicado hasta "
            f"<b>{_fmt_dt(_frontera)}</b> ya fue procesado por la IA. "
            f"Pendientes en cola del worker: {_falt_ca + _falt_li} "
            f"(compra ágil {_falt_ca} · licitaciones {_falt_li}). {_ult_txt}"
            f"</div>"
        )
    except Exception:  # noqa: BLE001
        pass

    if not total:
        return _layout(
            "Cola de revisión",
            "<h1>Cola de revisión</h1>" + estado_worker + filtros
            + "<div class=vacio>No hay clasificaciones pendientes con ese filtro. 🎉</div>",
            usuario=usuario,
        )

    cat = _catalogo()
    aviso = f"<div class=aviso>{_e(msg)}</div>" if msg else ""

    # Pre-cuenta filas por supergrupo para inyectar headers con el total.
    # Lo calculamos sobre TODO el conjunto filtrado (no solo la hoja), porque
    # el revisor quiere saber cuántas hay en cada categoría aunque vea una.
    grupo_counts: dict = {}
    if total and estado == "pendientes":
        try:
            for r in _query(
                "SELECT COUNT(*) n, "
                "CASE "
                "  WHEN pactivo_nuevo IS NOT NULL AND pactivo_nuevo<>'' THEN 'z1' "
                "  WHEN pactivo_sugerido='Adjunto' AND tabla_origen='compra_agil' THEN 'y1' "
                "  WHEN pactivo_sugerido='Adjunto' AND tabla_origen='Licitaciones_diarias' THEN 'y2' "
                "  WHEN interes_sugerido=1 AND metodo IN ('cruce_base','historico') THEN 'a1' "
                "  WHEN interes_sugerido=1 AND metodo='modelo_pactivo' THEN 'a2' "
                "  WHEN interes_sugerido=1 AND metodo='regla_diccionario' THEN 'a3' "
                "  WHEN interes_sugerido=1 AND metodo='claude' THEN 'a4' "
                "  WHEN interes_sugerido=0 AND metodo='descarte_item' THEN 'b1' "
                "  WHEN interes_sugerido=0 AND metodo='modelo_descarte' THEN 'b2' "
                "  WHEN interes_sugerido=0 AND metodo='conflicto_regla_modelo' THEN 'b3' "
                "  WHEN interes_sugerido=0 AND metodo='historico' THEN 'b4' "
                "  WHEN interes_sugerido=0 AND metodo='claude' THEN 'b5' "
                "  ELSE 'zz' END k "
                f"FROM clasificador_ia_log WHERE {where} GROUP BY k",
                tuple(args),
            ):
                grupo_counts[r["k"]] = r["n"]
        except Exception:  # noqa: BLE001
            pass

    # VISTA según usuario: el admin (quien afina) ve la técnica con badges de
    # confianza/vía/entrenamiento — le dice de dónde viene el error. El resto
    # (aprobadores) ve la vista LIMPIA: solo color + descripción + datos para
    # decidir. El orden por fecha de cierre ya viene del SQL (global, no por hoja).
    bloques = []
    grupo_actual = None
    for n, f in enumerate(filas):
        # Encabezado de supergrupo SOLO en la vista técnica (admin). Para
        # aprobadores el orden es por fecha de cierre, sin supergrupos.
        if es_admin:
            sg_key, sg_etiq = _supergrupo(f)
            if sg_key != grupo_actual:
                grupo_actual = sg_key
                short_key = sg_key.split("_")[0] if "_" in sg_key else sg_key
                n_grupo = grupo_counts.get(short_key, "")
                n_html = f" <span class=grupo-n>({n_grupo:,} en total)</span>" if n_grupo else ""
                bloques.append(f"<div class=grupo-hdr>{sg_etiq}{n_html}</div>")
        conf = float(f.get("confianza") or 0)
        interes = f.get("interes_sugerido")
        es_nuevo = bool((f.get("pactivo_nuevo") or "").strip())
        revisada = bool(f.get("revisado"))
        es_posible = f.get("metodo") == "modelo_marcas_posible"
        if interes == 0:
            tipo_cls, badge = "t-descarte", "<span class='badge b-desc'>DESCARTE sugerido</span>"
        elif es_nuevo:
            tipo_cls, badge = "t-nuevo", "<span class='badge b-nuevo'>PACTIVO NUEVO</span>"
        elif es_posible:
            tipo_cls, badge = "t-posible", "<span class='badge b-posible'>POSIBLEMENTE de interés (REVISAR)</span>"
        else:
            tipo_cls, badge = "", "<span class='badge b-int'>INTERÉS sugerido</span>"
        # Badges técnicos (vía cascada + entrenamiento del modelo) SOLO para el
        # admin que afina. Los aprobadores no los ven (distraen del flujo).
        via = ent = ""
        if es_admin:
            _met = f.get("metodo")
            via = f" · <span class='badge b-met'>vía {_e(_METODOS.get(_met, _met or '?'))}</span>"
            _m = _modelo_descarte()
            if _m is not None:
                pd = prob_descarte(_m, f.get("descripcion"))
                if pd >= 0.5:
                    ent = f" · <span class='badge b-ent-d'>entrenamiento: DESCARTE {pd:.2f}</span>"
                else:
                    ent = f" · <span class='badge b-ent-i'>entrenamiento: INTERÉS {1 - pd:.2f}</span>"

        _lic_txt = _e(num_lic.get((f["tabla_origen"], f["fila_id"])) or "—")
        _oinfo = info_origen.get((f["tabla_origen"], f["fila_id"]), {})
        _cierre = _oinfo.get("cierre")
        _cierre_txt = _fmt_dt(_cierre) if _cierre else "—"
        _pub = _oinfo.get("publicacion")
        _pub_txt = _fmt_dt(_pub) if _pub else "—"
        _dem = (_oinfo.get("demandante") or "").strip()
        _tit = (_oinfo.get("titulo") or "").strip()

        if es_admin:
            # Encabezado TÉCNICO con timestamps. Para revisadas suma quién y cuándo.
            meta_html = (
                f"<div class=meta>{badge} &nbsp; "
                f"<b>Licitación {_lic_txt}</b>"
                f" · {_e(f['tabla_origen'])} #{f['fila_id']} · "
                f"<span class='badge {'b-baja' if conf < 0.7 else 'b-alta'}'>confianza {conf:.2f}</span>"
                f"{via}{ent}"
                f"<div class=ts>clasificada {_fmt_dt(f.get('creado_en'))}"
                + (
                    f" · revisada {_fmt_dt(f.get('revisado_en'))} "
                    f"por <b>{_e(f.get('revisado_por') or '—')}</b>"
                    if revisada else ""
                )
                + "</div></div>"
            )
        else:
            # Encabezado LIMPIO de aprobación: N° licitación primero (izq→der),
            # luego fecha de cierre. Sin confianza/vía/entrenamiento. El color de
            # la fila (verde/rojo/naranja) ya comunica interés/descarte/nuevo.
            meta_html = (
                f"<div class='meta meta-aprob'>"
                f"<b class=ap-lic>N° {_lic_txt}</b>"
                f"<span class=ap-pub>publicada: {_pub_txt}</span>"
                f"<span class=ap-cierre>cierra: {_cierre_txt}</span>"
                + (f"<span class=ap-dem>{_e(_dem[:70])}</span>" if _dem else "")
                + (
                    f"<span class=ap-rev>revisada por {_e(f.get('revisado_por') or '—')}</span>"
                    if revisada else ""
                )
                + "</div>"
            )

        # FILA YA REVISADA: solo lectura, sin checkbox ni form. Muestra el
        # veredicto humano (aprobada / corregida) y sus notas para auditoría.
        if revisada:
            if f.get("feedback_correcto") == 1:
                veredicto_etiq = "<span class='badge b-int'>✓ APROBADA</span>"
                fp_disp = ""
                fn_disp = ""
            else:
                fp = (f.get("feedback_pactivo") or "").strip()
                fn = (f.get("feedback_notas") or "").strip()
                if fp:
                    veredicto_etiq = (
                        f"<span class='badge b-warn'>✏ CORREGIDA</span> "
                        f"→ <b>{_e(fp)}</b>"
                    )
                else:
                    veredicto_etiq = "<span class='badge b-desc'>✗ DESCARTADA</span>"
                fp_disp = fp
                fn_disp = fn
            # Lo que PROPUSO la IA + lo que DECIDIÓ el humano (lado a lado).
            _int_ia = "INTERÉS" if f.get("interes_sugerido") == 1 else "DESCARTE"
            _p_ia = _e(f.get("pactivo_sugerido") or "—")
            _c_ia = _e(f.get("composicion_sugerida") or "—")
            _pr_ia = _e(f.get("presentacion_sugerida") or "—")
            ia_prop = (
                f"<div class=ia-prop><b>IA propuso:</b> {_int_ia} · "
                f"pactivo <b>{_p_ia}</b> · comp {_c_ia} · pres {_pr_ia}</div>"
            )
            quien = _e(f.get("revisado_por") or "—")
            cuando = _fmt_dt(f.get("revisado_en"))
            humano = (
                f"<div class=hu-dec>{veredicto_etiq} · "
                f"<span class=hu-rev>por <b>{quien}</b> el {cuando}</span>"
                + (f" · <span class=hu-mot>motivo: {_e(fn_disp)}</span>" if fn_disp else "")
                + "</div>"
            )
            razon_html = f"<div class=razon>IA razón: {_e(f.get('razon'))}</div>" if es_admin else ""
            # COLOR según el VEREDICTO humano (no la propuesta IA): aprobada
            # respeta el color original; corregida → naranja; descartada → rojo.
            if f.get("feedback_correcto") == 1:
                tipo_cls_final = tipo_cls
            elif (f.get("feedback_pactivo") or "").strip():
                tipo_cls_final = "t-nuevo"  # corregida
            else:
                tipo_cls_final = "t-descarte"  # descartada
            # R4 — Form de edición de fila REVISADA. Permite cambiar
            # pactivo/comp/pres post-aprobación. Registra cada cambio en
            # clasificador_ia_log_revisiones (auditoría/timeline). Compartido
            # entre toda la cola — no requiere permisos especiales.
            pact_act = (f.get("feedback_pactivo") or f.get("pactivo_sugerido") or "")
            info_rev = cat.get(normalizar(pact_act))
            comps_rev = info_rev["comp"] if info_rev else []
            press_rev = info_rev["pres"] if info_rev else []
            # Composición/presentación actuales del legacy (si existen):
            comp_act = f.get("composicion_sugerida") or ""
            pres_act = f.get("presentacion_sugerida") or ""
            edicion_revisada = (
                f"<details class='reedicion'>"
                f"<summary style='cursor:pointer;padding:4px 0;color:#2f6fb0;"
                f"font-size:13px'>✎ editar pactivo / composición / presentación</summary>"
                f"<form method=post action='/revisadas/{f['id']}/reeditar' "
                f"style='display:flex;gap:8px;align-items:flex-end;flex-wrap:wrap;"
                f"padding:8px;background:#f5f8fc;border-radius:6px;margin-top:4px'>"
                f"<div><label style='font-size:11px;color:#6b7689'>pactivo</label>"
                f"<input type=text name=pactivo class=f-pactivo "
                f"value=\"{_e(pact_act)}\" style='width:240px'></div>"
                f"<div><label style='font-size:11px;color:#6b7689'>composición</label>"
                + _select("composicion", f"r{n}", "f-comp", comp_act, comps_rev)
                + "</div>"
                f"<div><label style='font-size:11px;color:#6b7689'>presentación</label>"
                + _select("presentacion", f"r{n}", "f-pres", pres_act, press_rev)
                + "</div>"
                f"<input type=text name=motivo class=motivo placeholder='motivo del cambio' "
                f"required style='flex:1;min-width:160px'>"
                f"<button type=submit class='ap-bajo' style='background:#2f6fb0;color:#fff;"
                f"border:0;padding:7px 14px;border-radius:5px;font-weight:600;cursor:pointer'>"
                f"actualizar</button>"
                f"</form></details>"
            )
            bloques.append(
                f"<div class='fila revisada {tipo_cls_final}'>"
                f"<div class=fila-head><div style='width:24px'></div>{meta_html}</div>"
                f"<div class=desc>{_e((f.get('descripcion') or '')[:300])}</div>"
                f"{ia_prop}{humano}{razon_html}{edicion_revisada}"
                f"</div>"
            )
            continue

        # FILA PENDIENTE: editable, con checkbox para procesarla en el lote.
        info = cat.get(normalizar(f.get("pactivo_sugerido") or ""))
        comps = info["comp"] if info else []
        press = info["pres"] if info else []

        aviso_nuevo = ""
        if es_nuevo:
            aviso_nuevo = (
                "<div class=nuevo-aviso>⚠ Claude no encontró este producto en el "
                f"catálogo y propone un pactivo NUEVO: <b>{_e(f['pactivo_nuevo'])}</b>. "
                "Si corresponde a un pactivo que ya existe, elígelo abajo y corrige; "
                "si de verdad es nuevo, descártalo o anótalo en el motivo para que "
                "se evalúe agregarlo al catálogo.</div>"
            )
        elif es_posible:
            aviso_nuevo = (
                f"<div class=nuevo-aviso style='background:#fff3a8;border-color:#c5a72b;"
                f"color:#7a5f06'>🟡 POSIBLEMENTE de interés — el modelo de marcas "
                f"sugiere <b>{_e(f.get('pactivo_sugerido') or '?')}</b> con "
                f"confianza media ({conf:.2f}). Revisar contexto y completar "
                f"pactivo + composición + presentación antes de aprobar; o "
                f"descartar si no corresponde.</div>"
            )

        # Línea de edición — MISMOS name= en ambas vistas (el POST a /revisar-hoja
        # es idéntico). Antes se ocultaba la edición para el admin cuando la fila
        # era descarte (vista técnica colapsada), pero eso impedía al admin
        # rescatar un descarte. Pedido del usuario 2026-06-02: admin DEBE poder
        # editar todo lo pendiente, igual que Carolina.
        linea_oculta = ""
        edicion = (
            f"<input type=hidden name=log_id value='{f['id']}'>"
            f"<div class='linea linea-edicion' data-row='{n}'{linea_oculta}>"
            f"<select name=decision class=decision data-row='{n}'>"
            "<option value=aprobar selected>Aprobar</option>"
            "<option value=corregir>Corregir</option>"
            "<option value=descartar>Descartar</option></select>"
            "<label>pactivo</label>"
            f"<span class=pac-wrap><input name=pactivo class=f-pactivo autocomplete=off "
            f"data-row='{n}' data-sug=\"{_e(f.get('pactivo_sugerido'))}\" "
            f"value=\"{_e(f.get('pactivo_sugerido'))}\">"
            f"<div class=pac-dd data-row='{n}'></div></span>"
            "<label>comp</label>"
            + _select("composicion", n, "f-comp", f.get("composicion_sugerida"), comps)
            + "<label>pres</label>"
            + _select("presentacion", n, "f-pres", f.get("presentacion_sugerida"), press)
            + "<input class=motivo name=motivo "
            "placeholder='motivo (obligatorio si corriges o descartas)'>"
            "</div>"
        )

        if es_admin:
            bloques.append(
                f"<div class='fila {tipo_cls}' data-row='{n}'>"
                f"<div class=fila-head>"
                f"<input type=checkbox class=marcar name=procesar value='{f['id']}' "
                f"data-row='{n}' checked>"
                f"{meta_html}</div>"
                f"<div class=desc>{_e((f.get('descripcion') or '')[:300])}</div>"
                f"<div class=razon>IA: {_e(f.get('razon'))}</div>"
                + aviso_nuevo + edicion + "</div>"
            )
        else:
            # Vista de aprobación: descripción protagonista, luego el TÍTULO del
            # tender, luego edición. Botón para tildar de esta fila hacia arriba.
            tit_html = (
                f"<div class=ap-titulo><span class=ap-tag>Título</span> {_e(_tit[:300])}</div>"
                if _tit else ""
            )
            bloques.append(
                f"<div class='fila fila-aprob {tipo_cls}' data-row='{n}'>"
                f"<div class=fila-head>"
                f"<input type=checkbox class=marcar name=procesar value='{f['id']}' "
                f"data-row='{n}' checked>"
                f"{meta_html}</div>"
                f"<div class='desc desc-aprob'>{_e((f.get('descripcion') or '')[:500])}</div>"
                + tit_html + aviso_nuevo + edicion
                + f"<button type=button class=ap-bajo onclick='aprobarDesde({n})'>"
                "✓ aprobar de aquí hacia arriba (lo ya revisado)</button>"
                + "</div>"
            )

    # datalist único con todos los pactivos del catálogo (autocompletado nativo)
    opts = "".join(
        f"<option value=\"{_e(v['nombre'])}\">"
        for v in sorted(cat.values(), key=lambda x: x["nombre"])
    )
    datalist = f"<datalist id=lista_pactivos>{opts}</datalist>"

    qs_extras = (f"tabla={tabla}" if tabla else "",
                 f"tipo={tipo}" if tipo else "",
                 f"metodo={metodo}" if metodo else "",
                 f"conf={conf}" if conf else "",
                 f"rango={rango}" if rango else "",
                 f"desde={desde}" if desde else "",
                 f"hasta={hasta}" if hasta else "",
                 f"estado={estado}" if estado != "pendientes" else "",
                 f"por_hoja={por_hoja}" if por_hoja != POR_HOJA_DEFAULT else "")
    qs_base = "&".join(p for p in qs_extras if p)
    qs_base = ("&" + qs_base) if qs_base else ""
    n_hojas = (total + por_hoja - 1) // por_hoja
    pag = "<div class=pag>"
    if hoja > 1:
        pag += f"<a href='/revision?hoja={hoja-1}{qs_base}'>« anterior</a>"
    pag += f" hoja {hoja} de {n_hojas} "
    if hoja < n_hojas:
        pag += f"<a href='/revision?hoja={hoja+1}{qs_base}'>siguiente »</a>"
    pag += "</div>"

    # Badge de "outbox pendiente" — lotes que el revisor aprobó pero NO se
    # sincronizaron a la BD (BD inaccesible). NADA se pierde porque están en
    # disco. El badge invita al revisor a clickear "Sincronizar".
    n_outbox = 0
    try:
        import sync_pendientes as _sp_count
        n_outbox = len(_sp_count.listar_pendientes())
    except Exception:  # noqa: BLE001
        pass
    badge_outbox = (
        f"<a class=outbox-bad href='/sincronizar' title='Hay {n_outbox} lote(s) "
        f"aprobado(s) que no se sincronizaron con clásico. Click para reintentar.'>"
        f"⚠ {n_outbox} pendiente(s) de sincronizar</a>"
        if n_outbox else ""
    )

    # Solo pendientes muestra form + botones de aprobar. Para revisadas/todas
    # la vista es de auditoría: lista de lectura con timestamps y veredicto.
    titulo_h1 = {
        "pendientes": f"Cola de revisión · {total} pendientes {badge_outbox}",
        "revisadas": f"Revisadas · {total} cerradas {badge_outbox}",
        "todas": f"Todas las clasificaciones · {total} {badge_outbox}",
    }[estado]
    if estado == "pendientes":
        # El nombre del logueado va al campo `nombre_clasificador` de cada fila
        # aprobada — antes era un input que el revisor escribía a mano.
        nombre_rev = (usuario or {}).get("name", "—") if usuario else "—"
        cuerpo = (
            f"<h1>{titulo_h1}</h1>{estado_worker}{aviso}{filtros}"
            f"<form method=post action='/revisar-hoja' id='formhoja'>"
            f"<input type=hidden name=hoja value='{hoja}'>"
            f"<input type=hidden name=tabla value='{_e(tabla)}'>"
            f"<input type=hidden name=tipo value='{_e(tipo)}'>"
            f"<input type=hidden name=metodo value='{_e(metodo)}'>"
            f"<input type=hidden name=conf value='{_e(conf)}'>"
            f"<input type=hidden name=rango value='{_e(rango)}'>"
            f"<input type=hidden name=desde value='{_e(desde)}'>"
            f"<input type=hidden name=hasta value='{_e(hasta)}'>"
            f"<input type=hidden name=por_hoja value='{por_hoja}'>"
            f"<input type=hidden name=busqueda value='{_e(busqueda)}'>"
            f"<input type=hidden name=licitacion value='{_e(licitacion)}'>"
            "<div class=barra>"
            f"<label>Revisor:</label><b>{_e(nombre_rev)}</b>"
            "<button type=button class=sec onclick='marcarTodos(true)'>Tildar todas</button>"
            "<button type=button class=sec onclick='marcarTodos(false)'>Destildar todas</button>"
            "<button type=submit>Aprobar <span id=cuenta>0</span> marcadas</button>"
            "<span style='font-size:13px;color:#6b7689'>Por default todas vienen "
            "tildadas; destildá solo las que dudás (quedan pendientes para revisar luego). "
            "Editar pactivo/comp/pres marca la fila como \"Corregir\" auto.</span>"
            "</div>"
            + "".join(bloques) + pag + "</form>" + datalist + _JS
        )
    else:
        # Auditoría: sin form ni JS — solo la lista con paginación.
        cuerpo = (
            f"<h1>{titulo_h1}</h1>{aviso}{filtros}"
            + "".join(bloques) + pag
        )
    return _layout("Cola de revisión", cuerpo, usuario=usuario)


# JS: selects dependientes (pactivo→comp/pres) + auto-cambio a "Corregir" al editar.
_JS = """<script>
function fila(n){return {
  dec:document.querySelector('select.decision[data-row=\"'+n+'\"]'),
  pac:document.querySelector('input.f-pactivo[data-row=\"'+n+'\"]'),
  com:document.querySelector('select.f-comp[data-row=\"'+n+'\"]'),
  pre:document.querySelector('select.f-pres[data-row=\"'+n+'\"]')};}
function llenar(sel,opciones,actual){
  var vals=[]; actual=(actual||'').trim();
  [actual,'Sin Cla'].concat(opciones||[]).forEach(function(v){
    v=(v||'').trim(); if(v&&vals.indexOf(v)<0)vals.push(v);});
  sel.innerHTML='';
  vals.forEach(function(v){
    var o=document.createElement('option'); o.value=v; o.textContent=v;
    if(v===actual)o.selected=true; sel.appendChild(o);});
}
function marcar(n){
  var e=fila(n); if(e.dec.value==='descartar')return;  // descarte manual: respetar
  var cambio = e.pac.value.trim()!==e.pac.dataset.sug
            || e.com.value!==e.com.dataset.sug
            || e.pre.value!==e.pre.dataset.sug;
  e.dec.value = cambio ? 'corregir' : 'aprobar';
}
function alCambiarPactivo(n){
  var e=fila(n); marcar(n);
  fetch('/api/catalogo?pactivo='+encodeURIComponent(e.pac.value.trim()))
    .then(function(r){return r.json();})
    .then(function(d){ llenar(e.com,d.comp,e.com.value); llenar(e.pre,d.pres,e.pre.value); })
    .catch(function(){});
}
document.querySelectorAll('.fila[data-row]').forEach(function(f){
  var n=f.dataset.row, e=fila(n);
  if(e.pac){e.pac.addEventListener('change',function(){alCambiarPactivo(n);});}
  if(e.com){e.com.addEventListener('change',function(){marcar(n);});}
  if(e.pre){e.pre.addEventListener('change',function(){marcar(n);});}
});

// COMBOBOX de pactivo: dropdown propio anclado DEBAJO del campo (el <datalist>
// nativo de Chrome ocupaba toda la pantalla). Filtra al teclear, con scroll.
var PACTIVOS = Array.prototype.map.call(
  document.querySelectorAll('#lista_pactivos option'), function(o){return o.value;});
function pacFiltrar(inp){
  var dd = inp.parentNode.querySelector('.pac-dd');
  if(!dd) return;
  var q = inp.value.trim().toLowerCase();
  var res = PACTIVOS.filter(function(p){return p.toLowerCase().indexOf(q) >= 0;}).slice(0, 60);
  dd.innerHTML = res.map(function(p){
    return '<div class=opt>' + p.replace(/&/g,'&amp;').replace(/</g,'&lt;') + '</div>';
  }).join('');
  dd.classList.add('open');
  Array.prototype.forEach.call(dd.querySelectorAll('.opt'), function(o){
    o.addEventListener('mousedown', function(ev){
      ev.preventDefault();           // antes del blur, para alcanzar a setear
      inp.value = o.textContent;
      dd.classList.remove('open');
      alCambiarPactivo(inp.dataset.row);  // recarga comp/pres + marca "corregir"
    });
  });
}
document.querySelectorAll('input.f-pactivo').forEach(function(inp){
  inp.addEventListener('focus', function(){pacFiltrar(inp);});
  inp.addEventListener('input', function(){pacFiltrar(inp);});
  inp.addEventListener('blur', function(){
    setTimeout(function(){
      var dd = inp.parentNode.querySelector('.pac-dd'); if(dd) dd.classList.remove('open');
    }, 180);
  });
});

// Checkbox por fila: tildada = se procesa (aprobar/corregir/descartar segun el
// select); destildada = la fila NO se incluye en el batch y queda pendiente
// para revisar luego. Por default vienen tildadas — es el patrón "aprobar
// todas menos las que destildo".
function actualizarCuenta(){
  var marcadas = document.querySelectorAll('.fila input.marcar:checked').length;
  var span = document.getElementById('cuenta');
  if(span){span.textContent = marcadas;}
}
function aplicarEstadoMarca(cb){
  var fila = cb.closest('.fila');
  if(!fila) return;
  if(cb.checked){fila.classList.remove('skip');}
  else{fila.classList.add('skip');}
}
function marcarTodos(estado){
  document.querySelectorAll('.fila input.marcar').forEach(function(cb){
    cb.checked = estado;
    aplicarEstadoMarca(cb);
  });
  actualizarCuenta();
}
// Vista de aprobación: tildar de esta fila HACIA ARRIBA (todo lo que ya revisé
// bajando queda aprobado; lo de abajo, que aún no vi, no se toca). Las filas
// vienen ordenadas por fecha de cierre más próxima.
function aprobarDesde(n){
  document.querySelectorAll('.fila[data-row]').forEach(function(f){
    if(parseInt(f.dataset.row,10) <= n){
      var cb=f.querySelector('input.marcar');
      if(cb){cb.checked=true; aplicarEstadoMarca(cb);}
    }
  });
  actualizarCuenta();
}
document.querySelectorAll('.fila input.marcar').forEach(function(cb){
  aplicarEstadoMarca(cb);
  cb.addEventListener('change', function(){
    aplicarEstadoMarca(cb); actualizarCuenta();
  });
});
// click sobre la descripción / cuerpo del card también togglea el checkbox
// (para destildar más rápido al ojear); excluye clicks sobre inputs/labels.
document.querySelectorAll('.fila').forEach(function(f){
  f.addEventListener('click', function(ev){
    if(ev.target.closest('input, select, label, button, .linea-edicion, .nuevo-aviso')) return;
    var cb = f.querySelector('input.marcar');
    if(!cb) return;
    cb.checked = !cb.checked;
    aplicarEstadoMarca(cb); actualizarCuenta();
  });
});
// expandir la línea de edición de los descartes al hacer doble-click sobre la fila
document.querySelectorAll('.fila').forEach(function(f){
  f.addEventListener('dblclick', function(ev){
    var linea = f.querySelector('.linea-edicion');
    if(linea && linea.hidden){ linea.hidden = false; ev.preventDefault(); }
  });
});
actualizarCuenta();

// R5 — REGLA DE ORO en el frontend: aprobar/corregir como interés exige los
// 3 campos (pact + comp + pres). Antes del submit, valida filas marcadas y
// bloquea con mensaje + scroll al primer problema. Espejo de la validacion
// del backend (sync_pendientes._aplicar_item -> "incompleto"); este JS es
// para que el usuario LO VEA antes de mandar el form.
function filasIncompletas(){
  var probs = [];
  document.querySelectorAll('.fila input.marcar:checked').forEach(function(cb){
    var f = cb.closest('.fila');
    if(!f) return;
    var n = f.dataset.row;
    var e = fila(n);
    if(!e.dec || !e.pac) return;
    if(e.dec.value === 'descartar') return;  // descarte explícito: OK
    // Si IA dijo descarte (clase t-descarte) y dec=aprobar (apruebo el descarte
    // tal cual), no exige campos. Usar classList en vez de .badge.b-desc para
    // que funcione también en vista aprobador (que no renderiza el badge).
    if(e.dec.value === 'aprobar' && f.classList.contains('t-descarte')) return;
    var p = (e.pac.value || '').trim();
    var c = (e.com.value || '').trim();
    var pr = (e.pre.value || '').trim();
    var falta = !p || !c || !pr;
    // 'Sin Cla' / 'Sin cla' se acepta como comodín legítimo (no es vacío).
    if(falta) probs.push(n);
  });
  return probs;
}
function actualizarBotonAprobar(){
  var probs = filasIncompletas();
  var btns = document.querySelectorAll('button.ap-bajo, button[type=submit][name=procesar], button#btn-procesar');
  // Limpiar outlines previos
  document.querySelectorAll('.fila').forEach(function(f){f.style.outline=''; });
  var btn_principal = document.querySelector('#btn-procesar') ||
                      document.querySelector('button[type=submit].btn-principal');
  if(probs.length){
    probs.forEach(function(n){
      var f = document.querySelector('.fila[data-row="' + n + '"]');
      if(f) f.style.outline = '2px solid #c0392b';
    });
    if(btn_principal){
      btn_principal.disabled = true;
      btn_principal.title = probs.length + ' fila(s) marcadas con campos vacíos — '
        + 'completa pact + comp + pres o cambia a "descartar"';
    }
  } else if(btn_principal) {
    btn_principal.disabled = false;
    btn_principal.title = '';
  }
}
// Validar al cambiar cualquier input de las filas marcadas
['change','input','blur'].forEach(function(ev){
  document.addEventListener(ev, function(e){
    if(!e.target.closest('.fila input.f-pactivo, .fila select.f-comp, '
       + '.fila select.f-pres, .fila select.decision, .fila input.marcar')) return;
    actualizarBotonAprobar();
  }, true);
});
// Validar al hacer submit del form principal
document.querySelectorAll('form[action*="revisar-hoja"]').forEach(function(form){
  form.addEventListener('submit', function(ev){
    var probs = filasIncompletas();
    if(probs.length){
      ev.preventDefault();
      // En vez de bloquear, ofrecer destildar las incompletas y seguir.
      // Caso típico: usuario quiere aprobar descartas que sí editó; la página
      // también trae filas interés-IA con campos vacíos checked por default que
      // no quería tocar — destildarlas las deja pendientes para otra ronda.
      var msg = '⚠ Hay ' + probs.length + ' fila(s) marcadas como interés '
                + 'sin pactivo + composición + presentación.\\n\\n'
                + 'OK = destildar las ' + probs.length + ' incompletas y '
                + 'aprobar el resto (quedan pendientes para revisar luego).\\n'
                + 'Cancelar = revisarlas manualmente (te llevo a la primera).';
      if(confirm(msg)){
        probs.forEach(function(n){
          var cb = document.querySelector('.fila[data-row="' + n + '"] input.marcar');
          if(cb){ cb.checked = false; aplicarEstadoMarca(cb); }
        });
        actualizarCuenta();
        form.submit();  // re-enviar sin re-disparar el handler
      } else {
        var primera = document.querySelector('.fila[data-row="' + probs[0] + '"]');
        if(primera) primera.scrollIntoView({behavior:'smooth', block:'center'});
      }
      return false;
    }
  });
});
// Validación inicial al cargar
actualizarBotonAprobar();
</script>"""


@app.post("/revisar-hoja")
def revisar_hoja(
    request: Request,
    hoja: int = Form(1),
    tabla: str = Form(""),
    tipo: str = Form(""),
    metodo: str = Form(""),
    conf: str = Form(""),
    rango: str = Form(""),
    desde: str = Form(""),
    hasta: str = Form(""),
    por_hoja: int = Form(POR_HOJA_DEFAULT),
    busqueda: str = Form(""),
    licitacion: str = Form(""),
    log_id: list[str] = Form([]),
    decision: list[str] = Form([]),
    pactivo: list[str] = Form([]),
    composicion: list[str] = Form([]),
    presentacion: list[str] = Form([]),
    motivo: list[str] = Form([]),
    procesar: list[str] = Form([]),
):
    # OUTBOX — el lote se PERSISTE en /app/pending/ ANTES de tocar la BD.
    # Si la BD cae, el JSON queda y un cron/botón lo reintenta. Garantía:
    # nada de lo que el revisor aprueba se pierde, aunque MySQL esté caído.
    u = usuario_actual(request)
    revisor = (u["name"] if u else "").strip()[:80] or "anónimo"
    set_procesar = set(procesar)
    saltadas = 0
    sin_motivo_v = 0

    # Construir el lote con todo lo necesario para reintentar (idempotente)
    items = []
    for i, lid in enumerate(log_id):
        if lid not in set_procesar:
            saltadas += 1
            continue
        dec = decision[i] if i < len(decision) else "aprobar"
        mot = (motivo[i] if i < len(motivo) else "").strip()
        if dec in ("corregir", "descartar") and not mot:
            sin_motivo_v += 1
            continue
        items.append({
            "log_id": lid,
            "decision": dec,
            "pactivo": (pactivo[i] if i < len(pactivo) else "").strip(),
            "composicion": (composicion[i] if i < len(composicion) else "").strip(),
            "presentacion": (presentacion[i] if i < len(presentacion) else "").strip(),
            "motivo": mot,
        })

    aplicadas = 0
    ya_revisadas = 0
    incompletas = 0
    pendiente = False
    if items:
        import json as _json
        import sync_pendientes as _sp
        lote = {
            "ts": datetime.now().isoformat(timespec="seconds"),
            "revisor": revisor,
            "items": items,
        }
        # PASO CRÍTICO: persistir a disco con write+rename atómico
        fp = _sp.guardar_pending(lote)
        try:
            aplic, ya, _, incompl = _sp.aplicar_lote(lote)
            aplicadas, ya_revisadas, incompletas = aplic, ya, incompl
            fp.unlink()  # éxito: borrar el JSON
        except Exception as exc:  # noqa: BLE001
            # JSON queda en pending. Reintenta cron + botón "Sincronizar".
            pendiente = True
            _err = str(exc)[:200]
            import logging
            logging.getLogger("revisar-hoja").warning(
                "Lote %s NO sincronizó (%s) — queda en %s para retry",
                lote["ts"], _err, fp,
            )

    msg = f"{aplicadas} fila(s) revisada(s)."
    if ya_revisadas:
        msg += f" {ya_revisadas} ya estaba(n) procesadas (idempotente)."
    if saltadas:
        msg += f" {saltadas} destildada(s) quedaron pendientes."
    if sin_motivo_v:
        msg += f" {sin_motivo_v} sin motivo obligatorio."
    if incompletas:
        msg += (f" ⚠ {incompletas} fila(s) marcada(s) interés SIN pactivo/comp/"
                f"presentación — completá los 3 campos antes de aprobar.")
    if pendiente:
        msg = (f"⚠ Lote ({len(items)} fila[s]) GUARDADO en outbox pero no se "
               f"sincronizó con clásico (BD inaccesible). Reintento automático "
               f"corriendo. NADA se perdió.")
    qs = f"/revision?hoja={hoja}&msg={msg}"
    for k, v in (("tabla", tabla), ("tipo", tipo), ("metodo", metodo),
                 ("conf", conf), ("rango", rango), ("desde", desde),
                 ("hasta", hasta), ("busqueda", busqueda),
                 ("licitacion", licitacion)):
        if v:
            qs += f"&{k}={v}"
    if por_hoja != POR_HOJA_DEFAULT:
        qs += f"&por_hoja={por_hoja}"
    return RedirectResponse(qs, status_code=303)


# ----------------------------------------------------- Sincronizar pendientes ---
@app.get("/sincronizar", response_class=HTMLResponse)
def sincronizar_pendientes(request: Request) -> str:
    """Reintenta aplicar los lotes que quedaron en /app/pending/ porque la BD
    estaba caída cuando el revisor los aprobó. NADA se pierde: el JSON queda
    en disco hasta que se aplica.

    También corre por cron cada 5 min (`sync_pendientes.py` en CLI mode), por
    lo que este botón es para forzar el reintento manual cuando el revisor
    ve la alerta y quiere asegurarse."""
    usuario = usuario_actual(request)
    import sync_pendientes as _sp
    pendientes_antes = len(_sp.listar_pendientes())
    r = _sp.procesar_pendientes()
    pendientes_despues = len(_sp.listar_pendientes())

    msg = (
        f"Sincronización ejecutada · {pendientes_antes} lote(s) pendiente(s) "
        f"al inicio · {r['lotes_ok']} aplicado(s) · {r['lotes_fallidos']} "
        f"siguen pendientes · {r['filas_aplicadas']} filas aplicadas · "
        f"{r['filas_ya_revisadas']} ya estaban procesadas"
    )
    cuerpo = (
        "<h1>Sincronización del outbox</h1>"
        f"<div class=aviso>{_e(msg)}</div>"
        + ("<h2>Errores</h2><ul>"
           + "".join(f"<li>{_e(e)}</li>" for e in r["errores"])
           + "</ul>" if r["errores"] else "")
        + f"<p>Pendientes ahora: <b>{pendientes_despues}</b></p>"
        "<p><a href='/revision'>← volver a /revision</a></p>"
    )
    return _layout("Sincronizar", cuerpo, usuario=usuario)


# ----------------------------------------------------- Estadísticas / BI ---
# Hitos manuales del sistema: cuando hicimos cada ajuste relevante. Se
# muestran como marcadores verticales en los gráficos para que en
# gerencia se vea "después de este cambio, la métrica X mejoró".
_HITOS = [
    ("2026-05-22", "Deploy inicial IA en producción"),
    ("2026-05-22", "Capa opciones comp/pres por pactivo"),
    ("2026-05-23", "Modelo de pactivo v1 (94.6%)"),
    ("2026-05-24", "Cruce Base ampliado (+ analisis_precios.Base)"),
    ("2026-05-25", "Modelo de pactivo umbral 0.40"),
    ("2026-05-25", "Modelo entrenado (descarte+pactivo) consolidado"),
    ("2026-05-26", "Catálogo activo (filtro clientes vivos)"),
    ("2026-05-26", "Umbral modelo_pactivo bajado a 0.30"),
    ("2026-05-26", "Match reglas usa solo descripción (no título)"),
    ("2026-05-26", "Outbox de aprobaciones (no se pierde nada)"),
    ("2026-05-27", "BLACKLIST Pharmatender interno"),
    ("2026-05-27", "Regla 'Medio de Contraste' en sistema"),
]


@app.get("/estadisticas", response_class=HTMLResponse)
def estadisticas(request: Request, tabla: str = "compra_agil",
                 dias: int = 14) -> str:
    """Dashboard con gráficos para presentar a gerencia:
    - clasificaciones por día (humano vs IA, interés vs descarte)
    - acierto IA vs humano por día
    - costo Claude acumulado (test + producción)
    - distribución por etapa de la cascada

    Los datos los carga JS desde `/estadisticas.json` (más rápido refresh)."""
    usuario = usuario_actual(request)
    cuerpo = f"""
<h1>Estadísticas IA · Pharmatender</h1>
<div class=cards style="margin-bottom:14px">
  <div class=card>
    <label style='font-size:11px;color:#6b7689'>Tabla</label>
    <select id=f_tabla onchange='cargar()'>
      <option value='compra_agil'{' selected' if tabla == 'compra_agil' else ''}>Compras ágiles</option>
      <option value='Licitaciones_diarias'{' selected' if tabla == 'Licitaciones_diarias' else ''}>Licitaciones</option>
    </select>
  </div>
  <div class=card>
    <label style='font-size:11px;color:#6b7689'>Días</label>
    <select id=f_dias onchange='cargar()'>
      <option value=7{' selected' if dias == 7 else ''}>7</option>
      <option value=14{' selected' if dias == 14 else ''}>14</option>
      <option value=30{' selected' if dias == 30 else ''}>30</option>
      <option value=60{' selected' if dias == 60 else ''}>60</option>
    </select>
  </div>
  <div class=card style='flex:2'>
    <div id=resumen style='font-size:13px;color:#6b7689'>Cargando…</div>
  </div>
</div>

<h2>1. Clasificaciones por día — humano vs IA</h2>
<p style='font-size:12px;color:#6b7689'>Verde = interés, rojo = descarte. Líneas continuas = humanos (gestor_licitaciones). Líneas punteadas = sistema IA. Marcadores verticales = hitos donde ajustamos el sistema.</p>
<div style='background:#fff;padding:16px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:18px'>
  <canvas id=chart_clasif height=110></canvas>
</div>

<h2>2. Acierto IA vs humano</h2>
<p style='font-size:12px;color:#6b7689'>% de filas donde la IA coincidió con la decisión del humano (interés o descarte). Se mide solo sobre filas que un humano ya revisó.</p>
<div style='background:#fff;padding:16px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:18px'>
  <canvas id=chart_acierto height=110></canvas>
</div>

<h2>3. Costo Claude acumulado</h2>
<p style='font-size:12px;color:#6b7689'>Suma acumulada por día. Azul = producción (lo que pagamos al clasificar las nuevas). Naranja = backtest (experimentación). Presupuesto mensual: USD ${config.budget_usd:.0f}.</p>
<div style='background:#fff;padding:16px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:18px'>
  <canvas id=chart_costo height=110></canvas>
</div>

<h2>4. Distribución por etapa de la cascada</h2>
<p style='font-size:12px;color:#6b7689'>Cada barra es un día. Las etapas baratas (cruce Base, histórico, modelos entrenados) resuelven el grueso; solo el residuo llega a Claude.</p>
<div style='background:#fff;padding:16px;border-radius:10px;box-shadow:0 1px 3px rgba(0,0,0,.08);margin-bottom:18px'>
  <canvas id=chart_metodo height=120></canvas>
</div>

<script src='https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js'></script>
<script>
let chs = {{}};
async function cargar() {{
  const tabla = document.getElementById('f_tabla').value;
  const dias = document.getElementById('f_dias').value;
  document.getElementById('resumen').textContent = 'Cargando…';
  const r = await fetch('/estadisticas.json?tabla='+tabla+'&dias='+dias);
  const d = await r.json();
  document.getElementById('resumen').innerHTML = d.resumen_html;
  // destroy charts previos
  Object.values(chs).forEach(c => c && c.destroy());
  chs = {{}};
  renderClasif(d); renderAcierto(d); renderCosto(d); renderMetodo(d);
}}

function hitosAnnot(hitos) {{
  // Líneas verticales sobre los días de hitos del sistema
  return hitos.map(h => ({{
    type: 'line', xMin: h.fecha, xMax: h.fecha,
    borderColor: '#16263d', borderWidth: 1, borderDash: [4, 4],
    label: {{ display: true, content: h.etiqueta, position: 'start',
              backgroundColor: '#16263d', color: '#fff', font: {{size: 10}},
              padding: 3, rotation: 90 }}
  }}));
}}

function renderClasif(d) {{
  chs.clasif = new Chart(document.getElementById('chart_clasif'), {{
    type: 'line',
    data: {{
      labels: d.dias,
      datasets: [
        {{label:'Humano · interés', data:d.humano_interes, borderColor:'#1b6b3a', backgroundColor:'#1b6b3a', tension:0.3, pointRadius:3}},
        {{label:'Humano · descarte', data:d.humano_descarte, borderColor:'#c0392b', backgroundColor:'#c0392b', tension:0.3, pointRadius:3}},
        {{label:'IA · interés', data:d.ia_interes, borderColor:'#28a76f', borderDash:[6,3], tension:0.3, pointRadius:0}},
        {{label:'IA · descarte', data:d.ia_descarte, borderColor:'#e57373', borderDash:[6,3], tension:0.3, pointRadius:0}},
      ]
    }},
    options: {{
      responsive: true, plugins: {{ legend: {{position:'bottom'}} }},
      scales: {{ y: {{beginAtZero:true, title:{{display:true,text:'filas/día'}}}} }}
    }}
  }});
}}

function renderAcierto(d) {{
  chs.acierto = new Chart(document.getElementById('chart_acierto'), {{
    type: 'line',
    data: {{
      labels: d.dias,
      datasets: [
        {{label:'Acierto INTERÉS', data:d.acierto_int, borderColor:'#2f6fb0', backgroundColor:'#2f6fb030', tension:0.3, pointRadius:3, fill:false}},
        {{label:'Acierto PACTIVO', data:d.acierto_pact, borderColor:'#d68910', backgroundColor:'#d6891020', tension:0.3, pointRadius:3, fill:false}},
      ]
    }},
    options: {{
      responsive: true, plugins: {{ legend: {{position:'bottom'}} }},
      scales: {{ y: {{min:50, max:100, ticks:{{callback:v=>v+'%'}}, title:{{display:true,text:'% acuerdo'}}}} }}
    }}
  }});
}}

function renderCosto(d) {{
  chs.costo = new Chart(document.getElementById('chart_costo'), {{
    type: 'line',
    data: {{
      labels: d.dias,
      datasets: [
        {{label:'Producción acumulado', data:d.costo_prod_acum, borderColor:'#2f6fb0', backgroundColor:'#2f6fb030', tension:0.1, fill:true}},
        {{label:'Backtest acumulado',   data:d.costo_test_acum, borderColor:'#d68910', backgroundColor:'#d6891030', tension:0.1, fill:true}},
      ]
    }},
    options: {{
      responsive: true, plugins: {{ legend: {{position:'bottom'}} }},
      scales: {{ y: {{beginAtZero:true, title:{{display:true,text:'USD acumulado'}}, ticks:{{callback:v=>'$'+v}}}} }}
    }}
  }});
}}

function renderMetodo(d) {{
  const palette = {{
    'cruce_base':'#28a76f', 'historico':'#1b6b3a', 'descarte_item':'#9aaab8',
    'modelo_descarte':'#5c7993', 'modelo_pactivo':'#2f6fb0',
    'conflicto_regla_modelo':'#c0392b', 'regla_diccionario':'#f0a830',
    'claude':'#7a45ad'
  }};
  const datasets = d.metodos.map(m => ({{
    label: m.metodo, data: m.data, backgroundColor: palette[m.metodo] || '#888'
  }}));
  chs.metodo = new Chart(document.getElementById('chart_metodo'), {{
    type: 'bar',
    data: {{ labels: d.dias, datasets: datasets }},
    options: {{
      responsive: true, plugins:{{legend:{{position:'bottom'}}}},
      scales: {{ x:{{stacked:true}}, y:{{stacked:true,title:{{display:true,text:'filas/día'}}}} }}
    }}
  }});
}}

cargar();
</script>
"""
    return _layout("Estadísticas", cuerpo, usuario=usuario)


@app.get("/estadisticas.json")
def estadisticas_json(tabla: str = "compra_agil", dias: int = 14):
    """JSON con todas las series para el dashboard. Una sola query batch."""
    if tabla not in TABLAS_VALIDAS:
        tabla = "compra_agil"
    dias = max(7, min(90, dias))

    # 1) Clasificaciones por día — HUMANO (en la tabla origen)
    sql_humano = (
        f"SELECT DATE(fecha_clasificacion) d, estado_gestor e, COUNT(*) n "
        f"FROM `{tabla}` "
        f"WHERE fecha_clasificacion >= NOW() - INTERVAL {dias} DAY "
        "AND nombre_clasificador IS NOT NULL "
        "AND nombre_clasificador NOT REGEXP '^(Bot|BOT|IA_|Bot Eliminado)' "
        "GROUP BY d, e"
    )
    humano = _query(sql_humano)

    # 2) Clasificaciones por día — IA (clasificador_ia_log)
    sql_ia = (
        "SELECT DATE(creado_en) d, interes_sugerido e, COUNT(*) n "
        "FROM clasificador_ia_log "
        f"WHERE tabla_origen = %s AND creado_en >= NOW() - INTERVAL {dias} DAY "
        "GROUP BY d, e"
    )
    ia = _query(sql_ia, (tabla,))

    # 3) Acierto IA vs humano (sobre filas comparables)
    sql_acierto = (
        "SELECT DATE(t.fecha_clasificacion) d, COUNT(*) n, "
        "SUM(log.interes_sugerido = t.estado_gestor) a_int, "
        "SUM(log.interes_sugerido=1 AND t.estado_gestor=1) ambos_int, "
        "SUM(log.interes_sugerido=1 AND t.estado_gestor=1 "
        "    AND log.pactivo_sugerido = t.pactivo) a_pact "
        f"FROM clasificador_ia_log log JOIN `{tabla}` t ON t.id = log.fila_id "
        f"WHERE log.tabla_origen = %s "
        f"AND t.fecha_clasificacion >= NOW() - INTERVAL {dias} DAY "
        "AND t.nombre_clasificador IS NOT NULL "
        "AND t.nombre_clasificador NOT REGEXP '^(Bot|BOT|IA_|Bot Eliminado)' "
        "GROUP BY d"
    )
    acierto = _query(sql_acierto, (tabla,))

    # 4) Costo Claude por día — test / producción
    sql_costo = (
        "SELECT DATE(creado_en) d, contexto, SUM(costo_usd) c "
        f"FROM clasificador_ia_costos "
        f"WHERE creado_en >= NOW() - INTERVAL {dias} DAY "
        "GROUP BY d, contexto"
    )
    costo = _query(sql_costo)

    # 5) Por método por día (stacked)
    sql_met = (
        "SELECT DATE(creado_en) d, metodo, COUNT(*) n "
        "FROM clasificador_ia_log "
        f"WHERE tabla_origen = %s AND creado_en >= NOW() - INTERVAL {dias} DAY "
        "GROUP BY d, metodo"
    )
    metodos_rows = _query(sql_met, (tabla,))

    # Construye lista de días continuos para que los charts no tengan huecos
    from datetime import timedelta
    hoy = datetime.now().date()
    dias_lista = [str(hoy - timedelta(days=i)) for i in range(dias - 1, -1, -1)]

    def map_por_dia(rows, key="d"):
        return {str(r[key]): r for r in rows}

    # Series humano
    h_int = {str(r["d"]): int(r["n"]) for r in humano if r.get("e") == 1}
    h_des = {str(r["d"]): int(r["n"]) for r in humano if r.get("e") == 0}
    # Series IA
    i_int = {str(r["d"]): int(r["n"]) for r in ia if r.get("e") == 1}
    i_des = {str(r["d"]): int(r["n"]) for r in ia if r.get("e") == 0}
    # Acierto
    acc_int = {}
    acc_pact = {}
    for r in acierto:
        d = str(r["d"])
        n = r.get("n") or 0
        if n:
            acc_int[d] = round((r.get("a_int") or 0) / n * 100, 1)
        ai = r.get("ambos_int") or 0
        if ai:
            acc_pact[d] = round((r.get("a_pact") or 0) / ai * 100, 1)
    # Costo acumulado (test y producción se grafican por separado, suma acumulada)
    costo_prod_dia = {str(r["d"]): float(r["c"]) for r in costo if r["contexto"] == "produccion"}
    costo_test_dia = {str(r["d"]): float(r["c"]) for r in costo if r["contexto"] == "test"}
    costo_prod_acum, costo_test_acum, acum_p, acum_t = [], [], 0.0, 0.0
    for d in dias_lista:
        acum_p += costo_prod_dia.get(d, 0.0)
        acum_t += costo_test_dia.get(d, 0.0)
        costo_prod_acum.append(round(acum_p, 3))
        costo_test_acum.append(round(acum_t, 3))
    # Métodos: armar estructura {metodo: {dia: count}} y devolver una serie por método
    metodos_set = sorted({r["metodo"] or "" for r in metodos_rows})
    met_dict = {m: {d: 0 for d in dias_lista} for m in metodos_set}
    for r in metodos_rows:
        m = r["metodo"] or ""
        d = str(r["d"])
        if d in met_dict[m]:
            met_dict[m][d] = int(r["n"])
    metodos_series = [
        {"metodo": m, "data": [met_dict[m][d] for d in dias_lista]}
        for m in metodos_set
    ]

    # Resumen humano-legible
    tot_humano = sum(h_int.values()) + sum(h_des.values())
    tot_ia = sum(i_int.values()) + sum(i_des.values())
    costo_total_prod = sum(costo_prod_dia.values())
    costo_total_test = sum(costo_test_dia.values())
    # Acierto promedio ponderado
    n_acc_int, n_total, n_pact, ambos = 0, 0, 0, 0
    for r in acierto:
        n_acc_int += r.get("a_int") or 0
        n_total += r.get("n") or 0
        n_pact += r.get("a_pact") or 0
        ambos += r.get("ambos_int") or 0
    acc_global_int = f"{n_acc_int/n_total*100:.1f}%" if n_total else "—"
    acc_global_pact = f"{n_pact/ambos*100:.1f}%" if ambos else "—"
    resumen_html = (
        f"<b>{tabla}</b> · últimos <b>{dias} días</b> &nbsp;|&nbsp; "
        f"Humano clasificó: <b>{tot_humano:,}</b> &nbsp; "
        f"IA clasificó: <b>{tot_ia:,}</b> &nbsp;|&nbsp; "
        f"Acierto interés: <b>{acc_global_int}</b> &nbsp; "
        f"Acierto pactivo: <b>{acc_global_pact}</b> &nbsp;|&nbsp; "
        f"Costo prod: <b>${costo_total_prod:.2f}</b> &nbsp; "
        f"backtest: <b>${costo_total_test:.2f}</b>"
    )

    return JSONResponse({
        "dias": dias_lista,
        "humano_interes": [h_int.get(d, 0) for d in dias_lista],
        "humano_descarte": [h_des.get(d, 0) for d in dias_lista],
        "ia_interes": [i_int.get(d, 0) for d in dias_lista],
        "ia_descarte": [i_des.get(d, 0) for d in dias_lista],
        "acierto_int": [acc_int.get(d) for d in dias_lista],
        "acierto_pact": [acc_pact.get(d) for d in dias_lista],
        "costo_prod_acum": costo_prod_acum,
        "costo_test_acum": costo_test_acum,
        "metodos": metodos_series,
        "hitos": [{"fecha": f, "etiqueta": e} for f, e in _HITOS],
        "resumen_html": resumen_html,
    })


# ----------------------------------------------------------------- Reglas ---
_VETO_FIRES_CACHE = {"ts": 0.0, "data": {}}


def _veto_fires(dias: int = 7) -> dict:
    """Cuántas veces disparó cada veto en los últimos `dias` (cuenta
    metodo='veto_<nombre>'). Mide los vetos de DESCARTE DURO (inicio_cascada y
    claude, que dejan metodo=veto_). Los por-rama (modelo/regla) ANULAN la
    predicción y la cascada sigue → no dejan metodo=veto_, no se miden así (su
    efecto se ve en el reporte agregado). Cache 10 min — query liviano (GROUP BY)."""
    import time as _t
    ahora = _t.time()
    c = _VETO_FIRES_CACHE
    if c["data"] and ahora - c["ts"] < 600:
        return c["data"]
    out: dict = {}
    try:
        # veto_aplicado cubre TODAS las ramas (incl. per-rama que antes eran
        # invisibles). fn = el veto descartó algo que el humano rescató
        # (interes=0, feedback_correcto=0, feedback_pactivo lleno) = over-firing.
        rows = _query(
            "SELECT veto_aplicado v, COUNT(*) n, "
            "SUM(interes_sugerido=0 AND feedback_correcto=0 AND feedback_pactivo IS NOT NULL "
            "AND feedback_pactivo<>'') fn "
            "FROM clasificador_ia_log "
            "WHERE veto_aplicado IS NOT NULL AND creado_en >= (NOW() - INTERVAL %s DAY) "
            "GROUP BY veto_aplicado", (dias,))
        for r in rows:
            out[r["v"]] = {"fires": r["n"], "fn": int(r["fn"] or 0)}
    except Exception:  # noqa: BLE001
        out = {}
    c["ts"], c["data"] = ahora, out
    return out


@app.get("/reglas", response_class=HTMLResponse)
def reglas(request: Request, msg: str = "") -> str:
    usuario = usuario_actual(request)
    try:
        items = _query(
            "SELECT id, tipo, texto, creado_por, creado_en, activa "
            "FROM clasificador_ia_reglas WHERE activa=1 "
            "ORDER BY tipo, creado_en DESC LIMIT 500"
        )
    except Exception as exc:  # noqa: BLE001
        return _layout("Error", f"<div class=vacio>{_e(exc)}</div>", usuario=usuario)

    reglas_ = [x for x in items if x["tipo"] == "regla"]
    corr = [x for x in items if x["tipo"] == "correccion"]

    # creado_en está guardado en UTC del container; lo mostramos en hora Chile
    # (UTC-4). Mismo offset que usamos en /revision.
    from datetime import timedelta as _td2
    _TZ_CL = _td2(hours=4)
    def _ts_chile(v):
        if not v:
            return ""
        if hasattr(v, "strftime"):
            return (v - _TZ_CL).strftime("%Y-%m-%d %H:%M")
        return str(v)[:16]
    def _dia_chile(v):
        if not v:
            return ""
        if hasattr(v, "strftime"):
            return (v - _TZ_CL).strftime("%Y-%m-%d")
        return str(v)[:10]

    def tabla(rows):
        """Reglas colapsables — cada una es un <details> con título extraído
        del texto (lo que va entre **...** al inicio) y cuerpo expandible."""
        if not rows:
            return "<div class=vacio>Sin registros.</div>"
        import re as _re
        out = []
        for r in rows:
            txt = (r["texto"] or "").strip()
            # Extraer título de los primeros **...**
            m = _re.match(r"\*\*([^*]{3,120})\*\*", txt)
            titulo = m.group(1).strip(" :—-") if m else (txt[:80] + "…")
            resto = txt[m.end():].strip(" :—-\n") if m else txt
            por = _e(r["creado_por"] or "")
            fecha = _e(_ts_chile(r["creado_en"]))
            # Cuerpo: convertir saltos de línea / negritas básicas a HTML
            cuerpo_html = _e(resto).replace("\n", "<br>")
            cuerpo_html = _re.sub(r"\*\*([^*]+?)\*\*", r"<b>\1</b>", cuerpo_html)
            out.append(
                f"<details class='regla-card'>"
                f"<summary><span class='regla-titulo'>{_e(titulo)}</span>"
                f"<span class='regla-meta'>{por} · {fecha}</span></summary>"
                f"<div class='regla-body'>{cuerpo_html}</div>"
                f"</details>"
            )
        return "<div class='regla-list'>" + "".join(out) + "</div>"

    def tabla_correcciones_por_dia(rows):
        """Agrupa correcciones por DATE(creado_en), día más reciente arriba.
        Cada día es un <details> colapsable con la cantidad en el resumen — para
        revisar/auditar sin tener que scrollear 100+ correcciones de un tirón."""
        if not rows:
            return "<div class=vacio>Sin registros.</div>"
        por_dia: dict = {}
        for r in rows:
            dia = _dia_chile(r["creado_en"])  # día agrupador ya en hora Chile
            por_dia.setdefault(dia, []).append(r)
        out = []
        for dia in sorted(por_dia.keys(), reverse=True):
            lst = por_dia[dia]
            filas = "".join(
                f"<tr><td>{_e(r['texto'])}</td><td>{_e(r['creado_por'])}</td>"
                f"<td>{_e(_ts_chile(r['creado_en'])[11:16])}</td></tr>"
                for r in lst
            )
            out.append(
                f"<details class=dia-grp open><summary>"
                f"<b>{dia}</b> · {len(lst)} corrección(es)"
                f"</summary>"
                f"<table><tr><th>Texto</th><th>Por</th><th>Hora</th></tr>"
                f"{filas}</table></details>"
            )
        return "".join(out)

    # Vetos (tipo='veto') — regex de la cascada, editables sin deploy.
    # NO se inyectan al prompt, cero costo extra. CRUD compartido con reglas
    # porque viven en la misma tabla; la sección lista los activos + form.
    try:
        vetos = _query(
            "SELECT id, nombre, regex_pattern, aplica_a, pactivo_filtro, "
            "texto, creado_por, creado_en, activa, protegido "
            "FROM clasificador_ia_reglas WHERE tipo='veto' "
            "ORDER BY activa DESC, protegido DESC, creado_en DESC LIMIT 200"
        )
    except Exception:
        vetos = []
    es_admin = _es_admin(usuario)
    fires = _veto_fires()  # {nombre: {fires, fn}} por veto (7d, vía veto_aplicado)

    def _badge_fires(r) -> str:
        st = fires.get(r["nombre"])
        if not r["activa"]:
            return ""
        if not st or not st["fires"]:
            return "<span class='veto-fires zero'>0 disparos (7d)</span>"
        b = f"<span class='veto-fires hit'>🔥 {st['fires']} disparos (7d)</span>"
        if st["fn"]:
            # FN = el veto descartó algo que el humano rescató → over-firing
            b += (f" <span class='veto-fires fn' title='El humano rescató "
                  f"{st['fn']} de las que este veto eliminó — revisar'>⚠️ {st['fn']} FN</span>")
        return b

    def tabla_vetos(rows):
        """Vetos compactos — cada uno una fila con nombre, dónde aplica,
        razón y botón. La regex completa va en details expandible."""
        if not rows:
            return "<div class=vacio>Sin vetos.</div>"
        out = ["<div class='veto-list'>"]
        for r in rows:
            estado_cls = "veto-on" if r["activa"] else "veto-off"
            estado_txt = "✓ activo" if r["activa"] else "✗ inactivo"
            prot = "🔒" if r["protegido"] else ""
            puede_tocar = r["activa"] and (es_admin or not r["protegido"]) or \
                          (not r["activa"] and (es_admin or not r["protegido"]))
            acc = ""
            if r["activa"] and (es_admin or not r["protegido"]):
                acc = (f"<form method=post action='/reglas/veto/{r['id']}/desactivar' "
                       f"style='display:inline'><button class='btn-mini' type=submit "
                       f"onclick=\"return confirm('Desactivar {_e(r['nombre'] or r['id'])}?')\">"
                       f"desactivar</button></form>")
            elif not r["activa"] and (es_admin or not r["protegido"]):
                acc = (f"<form method=post action='/reglas/veto/{r['id']}/activar' "
                       f"style='display:inline'><button class='btn-mini' type=submit>"
                       f"activar</button></form>")
            pact_f = f" · pact: <code>{_e(r['pactivo_filtro'])}</code>" if r['pactivo_filtro'] else ""
            out.append(
                f"<div class='veto-card {estado_cls}'>"
                f"<div class='veto-head'>"
                f"<span class='veto-nombre'>{prot} <b>{_e(r['nombre'] or '—')}</b></span>"
                f"<span class='veto-aplica'>{_e(r['aplica_a'] or 'inicio_cascada')}{pact_f}</span>"
                f"{_badge_fires(r)}"
                f"<span class='veto-estado'>{estado_txt}</span>"
                f"<span class='veto-acc'>{acc}</span>"
                f"</div>"
                f"<div class='veto-razon'>{_e(r['texto'] or '')}</div>"
                f"<details class='veto-regex'><summary>regex</summary>"
                f"<code>{_e(r['regex_pattern'] or '')}</code></details>"
                f"</div>"
            )
        out.append("</div>")
        return "".join(out)

    aviso = f"<div class=aviso>{_e(msg)}</div>" if msg else ""
    # CSS inline para mejorar la presentación. No depende de ui.py.
    css = """<style>
.regla-list { display: flex; flex-direction: column; gap: 8px; margin: 12px 0; }
.regla-card { border: 1px solid #d9e0ea; border-left: 4px solid #2f6fb0;
              border-radius: 7px; background: #fff; }
.regla-card summary { padding: 10px 14px; cursor: pointer; list-style: none;
                      display: flex; justify-content: space-between; gap: 12px;
                      flex-wrap: wrap; font-size: 14px; }
.regla-card summary::-webkit-details-marker { display: none; }
.regla-card[open] summary { border-bottom: 1px solid #eef1f4; background: #f5f8fc; }
.regla-titulo { font-weight: 600; color: #1d2330; flex: 1; min-width: 200px; }
.regla-meta { color: #6b7689; font-size: 12px; }
.regla-body { padding: 12px 16px 14px; font-size: 13.5px; line-height: 1.55;
              color: #2a3142; max-height: 460px; overflow-y: auto; }
.regla-body code { background: #eef1f4; padding: 1px 5px; border-radius: 3px;
                   font-size: 12px; }

.veto-list { display: flex; flex-direction: column; gap: 6px; margin: 12px 0; }
.veto-card { border: 1px solid #d9e0ea; border-radius: 6px; padding: 8px 12px;
             background: #fff; font-size: 13px; }
.veto-card.veto-on { border-left: 4px solid #1b6b3a; }
.veto-card.veto-off { border-left: 4px solid #c0392b; background: #faf2f2; opacity: 0.8; }
.veto-head { display: flex; gap: 14px; align-items: center; flex-wrap: wrap; }
.veto-nombre { font-size: 14px; min-width: 200px; }
.veto-aplica { color: #6b7689; font-size: 12px; flex: 1; }
.veto-aplica code { background: #eef1f4; padding: 1px 5px; border-radius: 3px;
                    font-size: 11px; }
.veto-estado { font-size: 12px; color: #6b7689; }
.veto-fires { font-size: 11.5px; padding: 1px 8px; border-radius: 10px; white-space: nowrap; }
.veto-fires.hit { background: #d9f0e1; color: #1b6b3a; font-weight: 600; }
.veto-fires.zero { background: #fdeccf; color: #7a4e09; }
.veto-fires.soft { background: #eef1f4; color: #6b7689; font-style: italic; }
.veto-fires.fn { background: #fbe4e1; color: #c0392b; font-weight: 700; }
.veto-razon { color: #44506a; font-size: 12.5px; margin-top: 4px;
              padding-left: 4px; border-left: 2px solid #d9e0ea; padding-left: 8px;
              padding-top: 4px; padding-bottom: 4px; }
.veto-regex { margin-top: 5px; font-size: 11px; }
.veto-regex summary { color: #6b7689; cursor: pointer; padding: 3px 0; }
.veto-regex code { background: #f6f7fa; padding: 6px 10px; border-radius: 4px;
                   display: block; word-break: break-all; font-size: 11px;
                   color: #4a4a4a; margin-top: 4px; }
.btn-mini { background: #fff; border: 1px solid #cdd5e0; border-radius: 5px;
            padding: 4px 12px; font-size: 12px; cursor: pointer;
            color: #2a3142; font-weight: 600; line-height: 1.2;
            font-family: inherit; }
.btn-mini:hover { background: #f3f6fa; border-color: #2f6fb0; color: #2f6fb0; }
.veto-card.veto-off .btn-mini { color: #1b6b3a; border-color: #1b6b3a; }
.veto-card.veto-off .btn-mini:hover { background: #e6f5ec; }

.seccion-titulo { display: flex; align-items: baseline; gap: 12px;
                  margin: 28px 0 8px; padding-bottom: 6px;
                  border-bottom: 2px solid #2f6fb0; }
.seccion-titulo h2 { margin: 0; font-size: 18px; }
.seccion-titulo .badge { font-size: 12px; padding: 2px 10px; border-radius: 14px;
                         background: #d9f0e1; color: #1b6b3a; }
.seccion-titulo .badge-warn { background: #fdeccf; color: #7a4e09; }
.seccion-desc { color: #6b7689; font-size: 13px; margin: 0 0 12px; }

form.alta-veto { display: grid; gap: 8px; max-width: 760px;
                 padding: 14px; background: #f5f8fc;
                 border: 1px solid #d9e0ea; border-radius: 7px;
                 margin-top: 10px; }
form.alta-veto input, form.alta-veto select { padding: 7px 10px;
            border: 1px solid #cdd5e0; border-radius: 5px; font-size: 13px; }
form.alta-veto button { padding: 8px 16px; background: #2f6fb0; color: #fff;
            border: 0; border-radius: 5px; font-weight: 600; cursor: pointer; }
form.alta-veto label { font-size: 12px; color: #6b7689; }
</style>"""
    n_vetos_act = sum(1 for v in vetos if v['activa'])
    cuerpo = (
        css
        + "<h1>Reglas, vetos y correcciones</h1>"
        + "<p class=seccion-desc>Tres tipos de feedback al sistema. Las "
        "<b>reglas</b> y <b>correcciones</b> van al prompt de Claude (tienen costo). "
        "Los <b>vetos</b> son regex en cascada (gratis).</p>"
        + aviso
        # === SECCIÓN 1: REGLAS DE NEGOCIO ===
        + f"<div class=seccion-titulo><h2>📋 Reglas de negocio</h2>"
        f"<span class=badge>{len(reglas_)} activas</span>"
        f"<span class='seccion-desc' style='margin:0'>· van al prompt de Claude (interpretativas)</span></div>"
        + "<form class=alta method=post action='/reglas' style='margin-bottom:6px'>"
        "<textarea name=texto placeholder='nueva regla de negocio para la IA' required></textarea>"
        "<button type=submit>Agregar regla</button></form>"
        + tabla(reglas_)
        # === SECCIÓN 2: VETOS EN CASCADA ===
        + f"<div class=seccion-titulo><h2>🛑 Vetos en cascada</h2>"
        f"<span class=badge>{n_vetos_act} activos</span>"
        f"<span class='seccion-desc' style='margin:0'>· regex sin costo, descartan antes de Claude</span></div>"
        + "<p class=seccion-desc>"
        "Patrones regex que la cascada aplica directamente. Descartan filas "
        "no farma (sensidiscos, stent, agua oxigenada, etc.) <b>sin gastar tokens</b>. "
        "Editables aquí, cambios aplican en ≤3 min sin deploy. Los <b>🔒 protegidos</b> "
        "solo el admin puede modificar.</p>"
        + tabla_vetos(vetos)
        + "<details style='margin-top:14px'><summary style='cursor:pointer;font-weight:600'>"
        "➕ Agregar nuevo veto</summary>"
        + "<form class=alta-veto method=post action='/reglas/veto'>"
        "<label>Nombre corto (sin espacios, ej. shampoo_no_farma)</label>"
        "<input name=nombre placeholder='shampoo_no_farma' required maxlength=100>"
        "<label>Regex (sintaxis Python, sin / al inicio/final)</label>"
        "<input name=regex_pattern placeholder='\\bshampoo\\b' required>"
        "<label>¿Dónde aplica? (todas activas — descartan antes de Claude o anulan la predicción de esa rama)</label>"
        "<select name=aplica_a>"
        "<option value='inicio_cascada'>inicio_cascada — descarta cualquier fila que matchee (antes de todo)</option>"
        "<option value='claude'>claude — anula la predicción de Claude (la rama de más FP)</option>"
        "<option value='modelo_pactivo'>modelo_pactivo — anula la predicción del modelo de pactivo</option>"
        "<option value='modelo_marcas'>modelo_marcas — anula la predicción del modelo de marcas</option>"
        "<option value='regla_diccionario'>regla_diccionario — anula el match simple del diccionario</option>"
        "</select>"
        "<label>Pactivo filtro (opcional — el veto solo dispara si la rama predijo este pactivo; vacío = cualquiera)</label>"
        "<input name=pactivo_filtro placeholder='ej. Cinta Adhesiva Médica'>"
        "<label>Razón / motivo (qué descarta y por qué)</label>"
        "<input name=texto placeholder='Ej: shampoo industrial, no es farma' required>"
        "<button type=submit>Agregar veto</button></form></details>"
        # === SECCIÓN 3: CORRECCIONES ===
        + f"<div class=seccion-titulo><h2>📝 Correcciones del equipo</h2>"
        f"<span class=badge>{len(corr)}</span>"
        f"<span class='seccion-desc' style='margin:0'>· errores ya corregidos, van al prompt con máxima prioridad</span></div>"
        + tabla_correcciones_por_dia(corr)
    )
    return _layout("Reglas", cuerpo, usuario=usuario)


@app.post("/reglas")
def agregar_regla(request: Request, texto: str = Form(...)):
    texto = texto.strip()
    if texto:
        u = usuario_actual(request)
        creado_por = ((u["name"] if u else "") or "").strip()[:80] or "anónimo"
        conn = conectar()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO clasificador_ia_reglas "
                    "(tipo, texto, creado_por, creado_en, activa) "
                    "VALUES ('regla',%s,%s,%s,1)",
                    (texto, creado_por, datetime.now()),
                )
            conn.commit()
        finally:
            conn.close()
    return RedirectResponse("/reglas?msg=Regla agregada.", status_code=303)


@app.post("/reglas/veto")
def agregar_veto(request: Request, nombre: str = Form(...),
                 regex_pattern: str = Form(...), aplica_a: str = Form("inicio_cascada"),
                 pactivo_filtro: str = Form(""), texto: str = Form(...)):
    """Agregar un veto nuevo. Valida que la regex compile antes de guardar."""
    import re as _re
    nombre = nombre.strip()[:100]
    regex_pattern = regex_pattern.strip()
    aplica_a = (aplica_a or "inicio_cascada").strip()
    pactivo_filtro = pactivo_filtro.strip() or None
    razon = texto.strip()
    if not (nombre and regex_pattern and razon):
        return RedirectResponse("/reglas?msg=Faltan campos obligatorios.", status_code=303)
    # Validar regex
    try:
        _re.compile(regex_pattern, _re.IGNORECASE)
    except _re.error as exc:
        return RedirectResponse(f"/reglas?msg=Regex inválida: {exc}", status_code=303)
    u = usuario_actual(request)
    creado_por = ((u["name"] if u else "") or "").strip()[:80] or "anónimo"
    conn = conectar()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO clasificador_ia_reglas "
                "(tipo, nombre, regex_pattern, aplica_a, pactivo_filtro, "
                "pactivo_asociado, texto, creado_por, creado_en, activa, protegido) "
                "VALUES ('veto',%s,%s,%s,%s,%s,%s,%s,%s,1,0)",
                (nombre, regex_pattern, aplica_a, pactivo_filtro,
                 pactivo_filtro, razon, creado_por, datetime.now()),
            )
        conn.commit()
    finally:
        conn.close()
    return RedirectResponse(f"/reglas?msg=Veto '{nombre}' agregado.", status_code=303)


@app.post("/reglas/veto/{id_}/desactivar")
def desactivar_veto(request: Request, id_: int):
    u = usuario_actual(request)
    es_admin = _es_admin(u)
    conn = conectar()
    try:
        with conn.cursor() as cur:
            # Si está protegido, solo admin puede tocar
            cur.execute("SELECT nombre, protegido FROM clasificador_ia_reglas WHERE id=%s AND tipo='veto'", (id_,))
            r = cur.fetchone()
            if not r:
                return RedirectResponse("/reglas?msg=Veto no existe.", status_code=303)
            if r["protegido"] and not es_admin:
                return RedirectResponse(f"/reglas?msg=Veto '{r['nombre']}' protegido — solo admin.", status_code=303)
            cur.execute("UPDATE clasificador_ia_reglas SET activa=0 WHERE id=%s", (id_,))
        conn.commit()
    finally:
        conn.close()
    return RedirectResponse("/reglas?msg=Veto desactivado.", status_code=303)


@app.post("/reglas/veto/{id_}/activar")
def activar_veto(request: Request, id_: int):
    u = usuario_actual(request)
    es_admin = _es_admin(u)
    conn = conectar()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT nombre, protegido FROM clasificador_ia_reglas WHERE id=%s AND tipo='veto'", (id_,))
            r = cur.fetchone()
            if not r:
                return RedirectResponse("/reglas?msg=Veto no existe.", status_code=303)
            if r["protegido"] and not es_admin:
                return RedirectResponse(f"/reglas?msg=Veto '{r['nombre']}' protegido — solo admin.", status_code=303)
            cur.execute("UPDATE clasificador_ia_reglas SET activa=1 WHERE id=%s", (id_,))
        conn.commit()
    finally:
        conn.close()
    return RedirectResponse("/reglas?msg=Veto activado.", status_code=303)


# --------------------------------------------- Pactivos extra (R8 2026-06-03) --
# Catálogo manual: permite agregar pactivos farma legítimos que NO están en Base
# ni los pide ningún cliente activo (caso 'Aztreonam-Avibactam'). Se SUMAN al
# catálogo activo. El worker detecta el cambio en el próximo ciclo (≤3 min)
# comparando version(MAX agregado_en, MAX desactivado_en, n filas, suma activos).
def _puede_pactivos_extra(usuario: dict | None) -> bool:
    """Solo admins + Carolina (clasificadora). Defensa en profundidad: el item
    del NAV se esconde, pero igual hay que blindar los handlers — alguien
    podría adivinar la URL."""
    email = ((usuario or {}).get("email") or "").strip().lower()
    return email in EMAILS_PACTIVOS_EXTRA


@app.get("/pactivos-extra", response_class=HTMLResponse)
def pactivos_extra_get(request: Request, msg: str = "") -> str:
    usuario = usuario_actual(request)
    if not _puede_pactivos_extra(usuario):
        return RedirectResponse("/revision", status_code=303)
    try:
        items = _query(
            "SELECT id, pactivo, agregado_por, agregado_en, activo, motivo, "
            "desactivado_por, desactivado_en "
            "FROM pactivos_extra ORDER BY activo DESC, agregado_en DESC LIMIT 500"
        )
    except Exception as exc:  # noqa: BLE001
        return _layout("Error", f"<div class=vacio>{_e(exc)}</div>", usuario=usuario)

    from datetime import timedelta as _td2
    _TZ_CL = _td2(hours=4)
    def _ts(v):
        if not v:
            return ""
        if hasattr(v, "strftime"):
            return (v - _TZ_CL).strftime("%Y-%m-%d %H:%M")
        return str(v)[:16]

    activos = [r for r in items if r["activo"]]
    inactivos = [r for r in items if not r["activo"]]

    # Vetos asociados a pactivos → para el 🔴 (un pactivo que se reintegra puede
    # tener un veto que lo descarta; el admin debe verlo). {pactivo_norm: [vetos]}
    vetos_pact: dict = {}
    try:
        for v in _query(
            "SELECT nombre, pactivo_asociado, activa FROM clasificador_ia_reglas "
            "WHERE tipo='veto' AND pactivo_asociado IS NOT NULL AND pactivo_asociado<>''"
        ):
            vetos_pact.setdefault((v["pactivo_asociado"] or "").strip().lower(), []).append(v)
    except Exception:  # noqa: BLE001
        vetos_pact = {}

    def _veto_badge(pact):
        vs = [v for v in vetos_pact.get((pact or "").strip().lower(), []) if v["activa"]]
        if not vs:
            return ""
        nombres = ", ".join(_e(v["nombre"]) for v in vs)
        return (f" <a href='/reglas' title='Veto ACTIVO que lo descarta: {nombres}' "
                f"style='color:#c0392b;font-weight:700;text-decoration:none'>🔴 veto</a>")

    def tabla(rows, mostrar_acciones: bool):
        if not rows:
            return "<div class=vacio>Sin registros.</div>"
        f = []
        for r in rows:
            acc = ""
            if mostrar_acciones:
                acc = (
                    f"<form method=post action='/pactivos-extra/{r['id']}/desactivar' "
                    f"style='display:inline'>"
                    f"<button type=submit onclick=\"return confirm('Desactivar "
                    f"{_e(r['pactivo'])}?')\">desactivar</button></form>"
                )
            inactivo_info = ""
            if not r["activo"] and r["desactivado_por"]:
                inactivo_info = (
                    f"<br><small>desactivado por {_e(r['desactivado_por'])} "
                    f"{_e(_ts(r['desactivado_en']))}</small>"
                )
            f.append(
                f"<tr><td><b>{_e(r['pactivo'])}</b>{_veto_badge(r['pactivo'])}{inactivo_info}</td>"
                f"<td>{_e(r['agregado_por'])}</td>"
                f"<td>{_e(_ts(r['agregado_en']))}</td>"
                f"<td><small>{_e((r['motivo'] or '')[:200])}</small></td>"
                f"<td>{acc}</td></tr>"
            )
        return ("<table><tr><th>Pactivo</th><th>Por</th><th>Fecha</th>"
                "<th>Motivo</th><th></th></tr>" + "".join(f) + "</table>")

    aviso = f"<div class=aviso>{_e(msg)}</div>" if msg else ""

    # Registro veto↔pactivo: ANTES de reintegrar un pactivo, el admin ve si tiene
    # un veto activo que lo descarta (caso "vuelve medio de contraste / alargador").
    activos_veto = {p: vs for p, vs in vetos_pact.items() if any(v["activa"] for v in vs)}
    if activos_veto:
        filas_rv = ""
        for vs in sorted(activos_veto.values(), key=lambda x: x[0]["pactivo_asociado"].lower()):
            pact = vs[0]["pactivo_asociado"]
            nombres = ", ".join(_e(v["nombre"]) for v in vs if v["activa"])
            filas_rv += (f"<tr><td><b>{_e(pact)}</b></td><td>{nombres}</td></tr>")
        registro_vetos = (
            "<div style='background:#fff4f2;border:1px solid #e0b4ad;border-left:5px solid "
            "#c0392b;border-radius:8px;padding:12px 16px;margin:14px 0'>"
            "<b style='color:#c0392b'>🔴 Pactivos con veto activo</b>"
            "<p style='font-size:13px;color:#7c3a30;margin:4px 0 10px'>Si vas a "
            "(re)integrar uno de estos, ojo: hay un veto que lo descarta. Revisalo/"
            "desactivalo en <a href='/reglas'>Reglas</a> primero.</p>"
            "<table><tr><th>Pactivo</th><th>Veto(s) activos</th></tr>"
            + filas_rv + "</table></div>"
        )
    else:
        registro_vetos = ""

    cuerpo = (
        "<h1>Pactivos extra — catálogo manual</h1>"
        + aviso
        + registro_vetos
        + "<p>Agrega pactivos farma legítimos que NO están en Base ni los pide "
        "ningún cliente activo. Se SUMAN al catálogo activo del clasificador. "
        "El worker los recoge en menos de 3 minutos.</p>"
        "<form class=alta method=post action='/pactivos-extra'>"
        "<input type=text name=pactivo placeholder='Nombre del pactivo (ej. Aztreonam-Avibactam)' required>"
        "<input type=text name=motivo placeholder='Motivo / contexto (opcional pero recomendado)'>"
        "<button type=submit>Agregar al catálogo</button></form>"
        f"<h2>Activos ({len(activos)})</h2>"
        + tabla(activos, mostrar_acciones=True)
        + f"<h2>Desactivados ({len(inactivos)})</h2>"
        + tabla(inactivos, mostrar_acciones=False)
    )
    return _layout("Pactivos extra", cuerpo, usuario=usuario)


@app.post("/pactivos-extra")
def pactivos_extra_post(
    request: Request,
    pactivo: str = Form(...),
    motivo: str = Form(""),
):
    if not _puede_pactivos_extra(usuario_actual(request)):
        return RedirectResponse("/revision", status_code=303)
    pact = (pactivo or "").strip()[:255]
    mot = (motivo or "").strip()[:5000] or None
    if not pact:
        return RedirectResponse("/pactivos-extra?msg=Pactivo vacío.", status_code=303)
    u = usuario_actual(request)
    por = ((u["name"] if u else "") or "").strip()[:80] or "anónimo"
    conn = conectar()
    try:
        with conn.cursor() as cur:
            # idempotente: si ya existía (activo o no), reactivar y actualizar
            cur.execute("SELECT id, activo FROM pactivos_extra WHERE pactivo=%s", (pact,))
            ex = cur.fetchone()
            if ex:
                cur.execute(
                    "UPDATE pactivos_extra SET activo=1, desactivado_en=NULL, "
                    "desactivado_por=NULL, motivo=COALESCE(%s, motivo) WHERE id=%s",
                    (mot, ex["id"]),
                )
                msg = (f"Pactivo '{pact}' reactivado." if not ex["activo"]
                       else f"Pactivo '{pact}' ya estaba activo (motivo actualizado).")
            else:
                cur.execute(
                    "INSERT INTO pactivos_extra (pactivo, agregado_por, agregado_en, "
                    "motivo, activo) VALUES (%s,%s,%s,%s,1)",
                    (pact, por, datetime.now(), mot),
                )
                msg = f"Pactivo '{pact}' agregado. Worker lo recogerá en <3 min."
        conn.commit()
        # Refrescar el catálogo del panel YA: así el pactivo aparece en el
        # autocompletado y en /api/catalogo sin esperar el TTL de 8h.
        try:
            _catalogo(forzar=True)
        except Exception:  # noqa: BLE001
            pass
    except Exception as exc:  # noqa: BLE001
        msg = f"Error: {exc}"
    finally:
        conn.close()
    return RedirectResponse(f"/pactivos-extra?msg={msg}", status_code=303)


@app.post("/pactivos-extra/{id_}/desactivar")
def pactivos_extra_desactivar(request: Request, id_: int):
    u = usuario_actual(request)
    if not _puede_pactivos_extra(u):
        return RedirectResponse("/revision", status_code=303)
    por = ((u["name"] if u else "") or "").strip()[:80] or "anónimo"
    conn = conectar()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE pactivos_extra SET activo=0, desactivado_en=%s, "
                "desactivado_por=%s WHERE id=%s",
                (datetime.now(), por, id_),
            )
        conn.commit()
        try:
            _catalogo(forzar=True)
        except Exception:  # noqa: BLE001
            pass
        msg = f"Pactivo id={id_} desactivado. Worker lo verá en <3 min."
    except Exception as exc:  # noqa: BLE001
        msg = f"Error: {exc}"
    finally:
        conn.close()
    return RedirectResponse(f"/pactivos-extra?msg={msg}", status_code=303)


@app.post("/revisadas/{log_id}/reeditar")
def reeditar_revisada(request: Request, log_id: int,
                      pactivo: str = Form(""),
                      composicion: str = Form(""),
                      presentacion: str = Form(""),
                      motivo: str = Form(...)):
    """R4 — re-editar una fila ya revisada. Actualiza compra_agil/Licitaciones_diarias
    + clasificador_ia_log + registra en clasificador_ia_log_revisiones (timeline).
    Cualquier usuario logueado puede usarlo.

    Regla de oro: si el estado es interés (estado_gestor=1), exige los 3 campos
    no vacíos. Si es descarte, los limpia.
    """
    u = usuario_actual(request)
    if not u:
        return RedirectResponse("/login", status_code=303)
    revisor = (u["name"] or "").strip()[:80] or "anónimo"
    pact = (pactivo or "").strip()
    comp = (composicion or "").strip()
    pres = (presentacion or "").strip()
    mot = (motivo or "").strip()
    if not mot:
        return RedirectResponse("/revision?estado=revisadas&msg=Motivo obligatorio.",
                                status_code=303)

    conn = conectar()
    try:
        with conn.cursor() as cur:
            # 1) Leer estado actual (origen + log)
            cur.execute(
                "SELECT tabla_origen, fila_id FROM clasificador_ia_log WHERE id=%s",
                (log_id,))
            r = cur.fetchone()
            if not r:
                return RedirectResponse("/revision?estado=revisadas&msg=Log no existe.",
                                        status_code=303)
            tabla_o = r["tabla_origen"]
            if tabla_o not in ("compra_agil", "Licitaciones_diarias"):
                return RedirectResponse("/revision?estado=revisadas&msg=Tabla inválida.",
                                        status_code=303)
            cur.execute(
                f"SELECT pactivo, composicion, presentacion, estado_gestor "
                f"FROM `{tabla_o}` WHERE id=%s", (r["fila_id"],))
            antes = cur.fetchone() or {}

            # 2) Decidir nuevo estado: si pactivo viene vacío → descarte; si
            # viene → interés con los 3 campos.
            if pact:
                if not (comp and pres):
                    return RedirectResponse(
                        f"/revision?estado=revisadas&msg=Para clasificar como "
                        f"interés se requiere pact + comp + pres (regla de oro).",
                        status_code=303)
                nuevo_estado = 1
                nueva_pact, nueva_comp, nueva_pres = pact, comp, pres
            else:
                nuevo_estado = 0
                nueva_pact, nueva_comp, nueva_pres = None, None, None

            # 3) Actualizar la tabla origen
            cur.execute(
                f"UPDATE `{tabla_o}` SET pactivo=%s, composicion=%s, "
                f"presentacion=%s, estado_gestor=%s, nombre_clasificador=%s, "
                f"fecha_clasificacion=%s WHERE id=%s",
                (nueva_pact, nueva_comp, nueva_pres, nuevo_estado, revisor,
                 datetime.now(), r["fila_id"])
            )
            # 4) Actualizar el log (feedback)
            cur.execute(
                "UPDATE clasificador_ia_log SET feedback_pactivo=%s, "
                "feedback_notas=CONCAT(IFNULL(feedback_notas,''), %s) "
                "WHERE id=%s",
                (nueva_pact, f"\n[reedit {datetime.now():%Y-%m-%d %H:%M}] {mot}",
                 log_id))
            # 5) Insertar en histórico
            cur.execute(
                "INSERT INTO clasificador_ia_log_revisiones "
                "(log_id, revisor, revisado_en, pactivo_antes, composicion_antes, "
                "presentacion_antes, estado_antes, pactivo_despues, composicion_despues, "
                "presentacion_despues, estado_despues, motivo) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (log_id, revisor, datetime.now(),
                 antes.get("pactivo"), antes.get("composicion"),
                 antes.get("presentacion"), antes.get("estado_gestor"),
                 nueva_pact, nueva_comp, nueva_pres, nuevo_estado, mot)
            )
        conn.commit()
    except Exception as exc:  # noqa: BLE001
        return RedirectResponse(f"/revision?estado=revisadas&msg=Error: {exc}",
                                status_code=303)
    finally:
        conn.close()
    return RedirectResponse(
        f"/revision?estado=revisadas&msg=Fila {log_id} actualizada por {revisor}.",
        status_code=303)


def _reporte_diario_html(rep: dict, dia: str) -> str:
    """Vista rica y explicada del reporte diario, desde el JSON estructurado.
    El foco: que se entienda QUÉ pasó y DÓNDE están los errores, con enlaces
    directos a /revision filtrado a esos casos (por fecha de revisión)."""
    p = rep.get("panel") or {}
    total = int(p.get("total") or 0)
    acuerdo = int(p.get("acuerdo") or 0)
    fp = int(p.get("fp") or 0)
    fn = int(p.get("fn") or 0)
    cambio = int(p.get("cambio_pact") or 0)
    errores = fp + fn + cambio
    pct_ok = (acuerdo / total * 100) if total else 0.0
    cst = rep.get("costo") or {}
    usd = float(cst.get("usd") or 0)
    catalogo = rep.get("pactivos_catalogo") or 0
    gen = rep.get("generado_en") or ""

    # Enlaces a los casos reales, filtrados por FECHA DE REVISIÓN del día.
    rd, rh = f"{dia}T00:00", f"{dia}T23:59"
    base = f"estado=revisadas&rev_desde={rd}&rev_hasta={rh}"
    link_todos = f"/revision?{base}&tipo_rev=corregidas"
    link_fp = f"/revision?{base}&tipo_rev=descartadas"
    link_fn = f"/revision?{base}&tipo_rev=corregidas&tipo=descarte"
    link_conf = f"/revision?{base}&tipo_rev=corregidas&tipo=interes"

    def _card(n, l, color=""):
        st = f" style='color:{color}'" if color else ""
        return f"<div class=card><div class=n{st}>{n}</div><div class=l>{_e(l)}</div></div>"

    def _tabla(titulo, expl, cols, filas, render):
        if not filas:
            return ""
        ths = "".join(f"<th>{_e(c)}</th>" for c in cols)
        trs = "".join("<tr>" + render(r) + "</tr>" for r in filas)
        return (f"<h2>{_e(titulo)}</h2>"
                f"<p class='rexpl'>{expl}</p>"
                f"<table><tr>{ths}</tr>{trs}</table>")

    css = """
<style>
.rexpl { font-size:13px; color:#6b7689; margin:2px 0 8px; }
.rep-banner { background:#eaf2fb; border:1px solid #c3d6ef; border-left:5px solid #2f6fb0;
  border-radius:8px; padding:12px 16px; font-size:13.5px; margin-bottom:16px; }
.rep-banner b { color:#1d2330; }
.err-box { background:#fff4f2; border:2px solid #e0b4ad; border-left:6px solid #c0392b;
  border-radius:10px; padding:14px 18px; margin:6px 0 22px; }
.err-box h2 { color:#c0392b; margin:0 0 4px; font-size:16px; }
.err-box .sub { font-size:13px; color:#7c3a30; margin-bottom:12px; }
.err-links { display:flex; flex-direction:column; gap:8px; }
.err-links a { display:flex; justify-content:space-between; align-items:center;
  background:#fff; border:1px solid #e0b4ad; border-radius:8px; padding:10px 14px;
  text-decoration:none; color:#1d2330; font-size:14px; }
.err-links a:hover { background:#fde9e6; border-color:#c0392b; }
.err-links a .cnt { font-weight:700; color:#c0392b; font-size:15px; }
.err-links a.todos { background:#c0392b; color:#fff; border-color:#a5281c; }
.err-links a.todos:hover { background:#a5281c; }
.err-links a.todos .cnt { color:#fff; }
td code, th code { background:#f3f6fa; padding:1px 5px; border-radius:4px; font-size:12.5px; }
.rep-md td:last-child { color:#44506a; font-size:12.5px; }
</style>
"""

    out = [css]
    out.append(
        f"<h1>📊 Reporte del día — {_e(dia)}</h1>"
        f"<div class=rep-banner>ℹ️ <b>Qué es esto:</b> resumen de lo que el "
        f"clasificador IA propuso ese día y <b>en qué se equivocó</b> (según lo "
        f"que el revisor humano corrigió en el panel). Para <b>ver y revisar los "
        f"errores</b>, usá los botones rojos de abajo. "
        f"<span style='color:#6b7689'>Generado {_e(gen[:19])} · catálogo {catalogo} pactivos.</span></div>"
    )

    # Tarjetas resumen
    out.append("<div class=cards>")
    out.append(_card(f"{total:,}".replace(",", "."), "revisadas en total"))
    out.append(_card(f"{errores:,}".replace(",", "."), "ERRORES (humano corrigió)", "#c0392b"))
    out.append(_card(f"{pct_ok:.1f}%", "acierto global"))
    out.append(_card(f"${usd:.2f}", "costo Claude del día"))
    out.append("</div>")

    # Bloque de ERRORES con enlaces
    out.append(
        "<div class=err-box>"
        "<h2>🔴 ¿Dónde están los errores?</h2>"
        "<div class=sub>Cada botón abre la cola <code>/revision</code> filtrada "
        "exactamente a esos casos (por fecha de revisión del día). Ahí podés ver "
        "la descripción completa y re-corregir si hace falta.</div>"
        "<div class=err-links>"
        f"<a class=todos href='{link_todos}'><span>▶ Ver TODOS los errores del día</span>"
        f"<span class=cnt>{errores}</span></a>"
        f"<a href='{link_fp}'><span>FP · la IA dijo <b>interés</b> y en realidad era descarte</span>"
        f"<span class=cnt>{fp}</span></a>"
        f"<a href='{link_fn}'><span>FN · la IA <b>descartó</b> algo que sí era de interés</span>"
        f"<span class=cnt>{fn}</span></a>"
        f"<a href='{link_conf}'><span>Confusión · acertó el interés pero <b>erró el pactivo</b></span>"
        f"<span class=cnt>{cambio}</span></a>"
        "</div></div>"
    )

    # FN causados por VETOS ese día (drill-down: del FN al veto culpable).
    # Live desde veto_aplicado — el veto descartó algo que el humano rescató.
    try:
        _ini = f"{dia} 04:00:00"  # día Chile en UTC (revisado_en es UTC, Chile=-4)
        from datetime import datetime as _dt, timedelta as _tdv
        _fin = (_dt.strptime(dia, "%Y-%m-%d") + _tdv(days=1)).strftime("%Y-%m-%d 04:00:00")
        fnv = _query(
            "SELECT veto_aplicado v, COUNT(*) n, LEFT(MIN(descripcion),60) ej "
            "FROM clasificador_ia_log WHERE veto_aplicado IS NOT NULL "
            "AND interes_sugerido=0 AND feedback_correcto=0 AND feedback_pactivo IS NOT NULL "
            "AND feedback_pactivo<>'' AND revisado_en >= %s AND revisado_en < %s "
            "GROUP BY veto_aplicado ORDER BY n DESC", (_ini, _fin))
    except Exception:  # noqa: BLE001
        fnv = []
    if fnv:
        filas_fnv = "".join(
            f"<tr><td><a href='/reglas'><code>{_e(r['v'])}</code></a></td>"
            f"<td>{r['n']}</td><td>{_e((r['ej'] or '')[:55])}</td></tr>" for r in fnv)
        out.append(
            "<div class=err-box style='border-left-color:#a5281c'>"
            "<h2>🚨 FN causados por VETOS (revisar el veto)</h2>"
            "<div class=sub>El veto eliminó algo que el humano rescató. Si un veto "
            "aparece acá seguido, está over-firing → ajustar/desactivar en "
            "<a href='/reglas'>Reglas</a>.</div>"
            "<table><tr><th>Veto</th><th>FN</th><th>ejemplo</th></tr>"
            + filas_fnv + "</table></div>"
        )

    # Detalle: tablas
    out.append(_tabla(
        "Top FP — interés de más", "La IA marcó interés pero el humano descartó. "
        "Si se repite un pactivo/vía, conviene un veto o ajuste de regla.",
        ["n", "IA propuso", "vía", "ejemplo"],
        rep.get("top_fp") or [],
        lambda r: f"<td>{r.get('n')}</td><td><code>{_e(r.get('p'))}</code></td>"
                  f"<td><code>{_e(r.get('metodo'))}</code></td><td>{_e((r.get('ej') or '')[:65])}</td>"))

    out.append(_tabla(
        "Top FN — descartes que sí servían", "La IA descartó algo que el humano "
        "rescató como de interés. Son los que más duele perder.",
        ["n", "pactivo humano", "vía descarte", "ejemplo"],
        rep.get("top_fn") or [],
        lambda r: f"<td>{r.get('n')}</td><td><code>{_e(r.get('hu'))}</code></td>"
                  f"<td><code>{_e(r.get('metodo'))}</code></td><td>{_e((r.get('ej') or '')[:65])}</td>"))

    out.append(_tabla(
        "Confusiones de pactivo", "Acertó que era de interés pero asignó el "
        "principio activo equivocado.",
        ["n", "IA dijo", "humano corrigió a", "ejemplo"],
        rep.get("top_confusiones") or [],
        lambda r: f"<td>{r.get('n')}</td><td><code>{_e(r.get('ia'))}</code></td>"
                  f"<td><code>{_e(r.get('hu'))}</code></td><td>{_e((r.get('ej') or '')[:65])}</td>"))

    crud = [r for r in (rep.get("candidatos_crud") or []) if (r.get("n") or 0) >= 2]
    out.append(_tabla(
        "Candidatos a catálogo nuevo", "Pactivos que el humano usó (≥2 veces en "
        "30 días) y NO están en el catálogo activo. Posibles altas vía CRUD pactivos_extra.",
        ["n", "pactivo", "ejemplo"],
        crud[:15],
        lambda r: f"<td>{r.get('n')}</td><td><code>{_e(r.get('p'))}</code></td>"
                  f"<td>{_e((r.get('ej') or '')[:65])}</td>"))

    out.append(_tabla(
        "Rendimiento por método (vía)", "Cuántas resolvió cada rama de la cascada "
        "y cuántas el humano marcó como error.",
        ["método", "n", "ok", "err"],
        rep.get("por_metodo") or [],
        lambda r: f"<td><code>{_e(r.get('metodo'))}</code></td><td>{r.get('n')}</td>"
                  f"<td>{r.get('ok')}</td><td>{r.get('err')}</td>"))

    out.append(f"<p style='margin-top:18px'><a href='/reportes?dia={_e(dia)}&fmt=md'>⬇ ver versión .md cruda</a></p>")
    return "\n".join(out)


@app.get("/reportes", response_class=HTMLResponse)
def reportes(request: Request, dia: str = "", tipo: str = "diario", fmt: str = "") -> str:
    """Reportes diarios (cron 22:00 UTC) + semanales (cron lunes 12:00 UTC)
    + corregidas (cron 14:30 UTC = 10:30 Chile, feedback humano)."""
    usuario = usuario_actual(request)
    if not usuario:
        return RedirectResponse("/login", status_code=303)
    from pathlib import Path
    rep_dir = Path("/app/reportes")
    rep_dir.mkdir(parents=True, exist_ok=True)
    diarios = sorted([f.stem.replace("reporte_", "")
                      for f in rep_dir.glob("reporte_*.md")], reverse=True)
    semanales = sorted([f.stem.replace("semanal_", "")
                        for f in rep_dir.glob("semanal_*.md")], reverse=True)
    corregidas = sorted([f.stem.replace("corregidas_", "")
                         for f in rep_dir.glob("corregidas_*.md")], reverse=True)
    if not diarios and not semanales and not corregidas:
        cuerpo = ("<h1>Reportes</h1>"
                  "<p>Aún no hay reportes. Diarios 18:00 Chile · "
                  "Corregidas 10:30 Chile · Semanales lunes 09:00 Chile.</p>")
        return _layout("Reportes", cuerpo, usuario=usuario)

    tabs = ("<p><a href='/reportes'>📊 diarios</a>  ·  "
            "<a href='/reportes?tipo=corregidas'>📋 corregidas</a>  ·  "
            "<a href='/reportes?tipo=semanal'>📅 semanales</a></p>")

    # ---- DIARIO: vista rica desde JSON (default = último día). fmt=md fuerza
    # el render markdown crudo (fallback / descarga visual).
    if tipo == "diario":
        if not dia and diarios:
            dia = diarios[0]
        nav = "  ·  ".join(
            (f"<a href='/reportes?dia={d}'>{d}</a>" if d != dia else f"<b>{d}</b>")
            for d in diarios[:14])
        json_p = rep_dir / f"reporte_{dia}.json"
        if dia and fmt != "md" and json_p.exists():
            import json as _json
            try:
                rep = _json.loads(json_p.read_text())
                cuerpo = f"{tabs}<p>{nav}</p>" + _reporte_diario_html(rep, dia)
                return _layout("Reportes", cuerpo, usuario=usuario)
            except Exception:  # noqa: BLE001 — si el JSON falla, cae al markdown
                pass

    # Decidir qué archivo .md abrir (semanal/corregidas, o diario con fmt=md)
    archivo = None
    if dia:
        if tipo == "semanal":
            p = rep_dir / f"semanal_{dia}.md"
        elif tipo == "corregidas":
            p = rep_dir / f"corregidas_{dia}.md"
        else:
            p = rep_dir / f"reporte_{dia}.md"
        if p.exists(): archivo = p
    if archivo:
        contenido = archivo.read_text()
        # Render simple de Markdown a HTML — limitado pero suficiente
        import re as _re
        html = _e(contenido)
        # headers
        html = _re.sub(r"^# (.+)$", r"<h1>\1</h1>", html, flags=_re.M)
        html = _re.sub(r"^## (.+)$", r"<h2>\1</h2>", html, flags=_re.M)
        # tablas: convertir | a HTML
        out_lines = []
        en_tabla = False
        for ln in html.split("\n"):
            if ln.startswith("| ") and "|" in ln[2:]:
                cells = [c.strip() for c in ln.strip().strip("|").split("|")]
                if all(_re.match(r"^-+:?$|^:?-+:?$", c) for c in cells):
                    continue  # separador
                tag = "th" if not en_tabla else "td"
                if not en_tabla:
                    out_lines.append("<table>")
                    en_tabla = True
                out_lines.append("<tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>")
            else:
                if en_tabla:
                    out_lines.append("</table>")
                    en_tabla = False
                out_lines.append(ln)
        if en_tabla: out_lines.append("</table>")
        html = "\n".join(out_lines).replace("`", "")
        # Nav según tipo
        tabs = ("<p><a href='/reportes'>📊 diarios</a>  ·  "
                "<a href='/reportes?tipo=corregidas'>📋 corregidas</a>  ·  "
                "<a href='/reportes?tipo=semanal'>📅 semanales</a></p>")
        if tipo == "semanal":
            nav = "  ·  ".join(
                (f"<a href='/reportes?dia={d}&tipo=semanal'>{d}</a>" if d != dia else f"<b>semana {d}</b>")
                for d in semanales[:12])
        elif tipo == "corregidas":
            nav = "  ·  ".join(
                (f"<a href='/reportes?dia={d}&tipo=corregidas'>{d}</a>" if d != dia else f"<b>{d}</b>")
                for d in corregidas[:14])
        else:
            nav = "  ·  ".join(
                (f"<a href='/reportes?dia={d}'>{d}</a>" if d != dia else f"<b>{d}</b>")
                for d in diarios[:14])
        cuerpo = (f"{tabs}<p>{nav}</p>"
                  f"<div class='reporte'>{html}</div>")
    elif tipo == "semanal":
        items = "".join(f"<li><a href='/reportes?dia={d}&tipo=semanal'>semana del {d}</a></li>"
                        for d in semanales[:30])
        cuerpo = (f"<h1>Reportes semanales</h1>"
                  f"<p><a href='/reportes'>📊 diarios</a>  ·  "
                  f"<a href='/reportes?tipo=corregidas'>📋 corregidas</a></p>"
                  f"<ul>{items or '<li><em>sin reportes semanales aún</em></li>'}</ul>")
    elif tipo == "corregidas":
        items = "".join(f"<li><a href='/reportes?dia={d}&tipo=corregidas'>{d}</a></li>"
                        for d in corregidas[:30])
        cuerpo = (f"<h1>📋 Reportes de CORREGIDAS</h1>"
                  f"<p><em>Feedback humano del día anterior: rebajes, rescates, "
                  f"correcciones de pactivo y efectividad de vetos.</em></p>"
                  f"<p><a href='/reportes'>📊 diarios</a>  ·  "
                  f"<a href='/reportes?tipo=semanal'>📅 semanales</a></p>"
                  f"<ul>{items or '<li><em>sin reportes corregidas aún (cron 10:30 Chile)</em></li>'}</ul>")
    else:
        items = "".join(f"<li><a href='/reportes?dia={d}'>{d}</a></li>"
                        for d in diarios[:30])
        cuerpo = (f"<h1>Reportes</h1>"
                  f"<p><a href='/reportes?tipo=corregidas'>📋 corregidas</a>  ·  "
                  f"<a href='/reportes?tipo=semanal'>📅 semanales</a></p>"
                  f"<h2>Diarios</h2><ul>{items}</ul>")
    return _layout("Reportes", cuerpo, usuario=usuario)


@app.get("/salud")
def salud():
    try:
        _query("SELECT 1")
        return {"estado": "ok"}
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"estado": "error", "detalle": str(exc)}, status_code=500)
